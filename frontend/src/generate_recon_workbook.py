#!/usr/bin/env python3
"""
Bank Reconciliation Workbook Generator
Production-grade Excel workbook for treasury reconciliation
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from copy import copy

# ─── Color Palette ───
class P:
    INK      = "0F1117"
    WHITE    = "FFFFFF"
    SLATE    = "3D4560"
    MID      = "6B7694"
    LIGHT    = "A8B0C8"
    RULE     = "DDE2EF"
    PALE     = "F0F2F8"
    CREAM    = "F7F8FB"
    AZURE    = "1A3A7C"
    AZURE2   = "2651A8"
    AZURE4   = "D4DFF5"
    EMERALD  = "0D5C3A"
    EMERALD3 = "D0F0E4"
    CRIMSON  = "C0202E"
    CRIMSON3 = "FCE8EA"
    AMBER    = "D46A00"
    AMBER3   = "FFF3E0"
    BG       = "F7F8FB"
    HEADER_BG = "0F1117"

# ─── Style helpers ───
thin_border = Border(
    left=Side(style='thin', color=P.RULE),
    right=Side(style='thin', color=P.RULE),
    top=Side(style='thin', color=P.RULE),
    bottom=Side(style='thin', color=P.RULE),
)
header_font = Font(name='Calibri', size=10, bold=True, color=P.WHITE)
header_fill = PatternFill(start_color=P.INK, end_color=P.INK, fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
data_font = Font(name='Calibri', size=10, color=P.SLATE)
data_align_r = Alignment(horizontal='right', vertical='center')
data_align_c = Alignment(horizontal='center', vertical='center')
data_align_l = Alignment(horizontal='left', vertical='center')
mono_font = Font(name='Consolas', size=10, color=P.SLATE)
title_font = Font(name='Calibri', size=18, bold=True, color=P.AZURE)
subtitle_font = Font(name='Calibri', size=12, color=P.MID)
kpi_val_font = Font(name='Consolas', size=22, bold=True, color=P.INK)
kpi_label_font = Font(name='Calibri', size=9, color=P.MID)
section_font = Font(name='Calibri', size=13, bold=True, color=P.AZURE)

SAR_FMT = '#,##0.00 "SAR"'
SAR_FMT_0 = '#,##0 "SAR"'
PCT_FMT = '0.0%'
DATE_FMT = 'YYYY-MM-DD'
NUM_FMT = '#,##0.00'

def style_header_row(ws, row, max_col, height=30):
    ws.row_dimensions[row].height = height
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

def style_data_area(ws, start_row, end_row, max_col, alt_fill=True):
    pale = PatternFill(start_color=P.PALE, end_color=P.PALE, fill_type='solid')
    white = PatternFill(start_color=P.WHITE, end_color=P.WHITE, fill_type='solid')
    for r in range(start_row, end_row + 1):
        fill = pale if (r - start_row) % 2 == 1 and alt_fill else white
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = data_font
            cell.border = thin_border
            cell.fill = fill
            cell.alignment = data_align_r

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def add_table(ws, ref, name, style="TableStyleMedium2"):
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name=style, showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(tbl)

DATA_ROWS = 1000  # pre-fill formula rows

wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════
# 1. README SHEET
# ═══════════════════════════════════════════════════════════════
ws_readme = wb.active
ws_readme.title = "README"
ws_readme.sheet_properties.tabColor = P.AZURE

ws_readme.sheet_view.showGridLines = False
set_col_widths(ws_readme, [3, 80])

readme_lines = [
    ("نظام مطابقة الحسابات البنكية", title_font),
    ("BANK RECONCILIATION SYSTEM", Font(name='Consolas', size=10, color=P.LIGHT)),
    ("", None),
    ("الغرض من هذا الملف", section_font),
    ("يقوم هذا الملف بمطابقة كشف الحساب البنكي مع قيود دفتر الأستاذ العام تلقائياً.", data_font),
    ("يتم احتساب حالة كل قيد: مطابق، قريب من المطابقة، أو غير مطابق.", data_font),
    ("", None),
    ("خطوات الاستخدام", section_font),
    ("1. افتح ورقة SETTINGS واضبط إعدادات التسامح والشركة.", data_font),
    ("2. الصق بيانات كشف البنك في ورقة BANK_RAW (بدون عناوين الأعمدة).", data_font),
    ("3. الصق بيانات الأستاذ العام في ورقة GL_RAW (بدون عناوين الأعمدة).", data_font),
    ("4. ستقوم أوراق BANK_CLEAN و GL_CLEAN بتنظيف البيانات تلقائياً.", data_font),
    ("5. ورقة MATCH_ENGINE تنفذ المطابقة تلقائياً.", data_font),
    ("6. راجع النتائج في RECON_RESULTS.", data_font),
    ("7. راجع لوحة المعلومات في DASHBOARD.", data_font),
    ("8. راجع الاستثناءات في EXCEPTIONS.", data_font),
    ("9. استخدم EXPORT_READY لتصدير النتائج.", data_font),
    ("", None),
    ("إعدادات التسامح", section_font),
    ("فرق المبلغ: الحد الأقصى المسموح للفرق بين مبلغ GL ومبلغ البنك (بالريال).", data_font),
    ("فرق الأيام: الحد الأقصى المسموح للفرق بين تاريخ GL وتاريخ البنك.", data_font),
    ("حد المطابقة القريبة: الحد الأدنى لنسبة التطابق لاعتبار القيد قريباً.", data_font),
    ("", None),
    ("حالات المطابقة", section_font),
    ("Matched (مطابق): الفرق في المبلغ والتاريخ ضمن حدود التسامح.", data_font),
    ("Near Match (قريب): خارج حدود التسامح لكن ضمن حد المطابقة القريبة.", data_font),
    ("Unmatched GL (GL غير مطابق): قيد أستاذ عام بدون حركة بنكية مقابلة.", data_font),
    ("Unmatched Bank (بنك غير مطابق): حركة بنكية بدون قيد أستاذ عام مقابل.", data_font),
    ("", None),
    ("ملاحظات مهمة", section_font),
    ("لا تعدل أوراق CLEAN أو MATCH_ENGINE أو RECON_RESULTS — تحتوي على صيغ.", data_font),
    ("عدل فقط: SETTINGS و BANK_RAW و GL_RAW.", data_font),
    ("الملف مصمم لـ Microsoft Excel 365.", data_font),
]

for i, (text, font) in enumerate(readme_lines, 2):
    cell = ws_readme.cell(row=i, column=2, value=text)
    if font:
        cell.font = font
    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

ws_readme.protection.sheet = True
ws_readme.protection.password = ""

# ═══════════════════════════════════════════════════════════════
# 2. SETTINGS SHEET
# ═══════════════════════════════════════════════════════════════
ws_set = wb.create_sheet("SETTINGS")
ws_set.sheet_properties.tabColor = P.AMBER
ws_set.sheet_view.showGridLines = False
set_col_widths(ws_set, [3, 35, 25, 35])

settings_data = [
    (2, "إعدادات النظام — SYSTEM SETTINGS", None, None, section_font),
    (4, "الإعداد", "القيمة", "الوصف", header_font),
    (5, "فرق المبلغ المسموح (SAR)", 1, "Amount Tolerance — الحد الأقصى للفرق بالريال", None),
    (6, "فرق الأيام المسموح", 1, "Date Tolerance — الحد الأقصى للفرق بالأيام", None),
    (7, "حد المطابقة القريبة (%)", 50, "Near Match Threshold — أقل نسبة تطابق للمطابقة القريبة", None),
    (8, "فترة التقرير", "2025-Q1", "Reporting Period", None),
    (10, "معلومات الشركة — COMPANY INFO", None, None, section_font),
    (12, "الإعداد", "القيمة", "الوصف", header_font),
    (13, "اسم الشركة", "شركة المثال التجارية", "Company Name", None),
    (14, "اسم الحساب البنكي", "الحساب الجاري الرئيسي", "Bank Account Name", None),
    (15, "رقم الحساب البنكي", "SA0380000000608010167519", "Bank Account Number (IBAN)", None),
    (16, "العملة", "SAR", "Currency", None),
    (17, "اسم البنك", "البنك الأهلي السعودي", "Bank Name", None),
]

for row_num, c1, c2, c3, fnt in settings_data:
    ws_set.cell(row=row_num, column=2, value=c1)
    if c2 is not None:
        ws_set.cell(row=row_num, column=3, value=c2)
    if c3 is not None:
        ws_set.cell(row=row_num, column=4, value=c3)
    if fnt == header_font:
        for c in range(2, 5):
            cell = ws_set.cell(row=row_num, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
    elif fnt == section_font:
        ws_set.cell(row=row_num, column=2).font = section_font
    else:
        for c in range(2, 5):
            cell = ws_set.cell(row=row_num, column=c)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = data_align_r

# Highlight value cells
val_fill = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type='solid')
for r in [5, 6, 7, 8, 13, 14, 15, 16, 17]:
    ws_set.cell(row=r, column=3).fill = val_fill
    ws_set.cell(row=r, column=3).font = Font(name='Consolas', size=11, bold=True, color=P.AZURE)

# Named Ranges
from openpyxl.workbook.defined_name import DefinedName
named = {
    "Tol_Amount":   "SETTINGS!$C$5",
    "Tol_Days":     "SETTINGS!$C$6",
    "Tol_NearPct":  "SETTINGS!$C$7",
    "Rpt_Period":   "SETTINGS!$C$8",
    "Co_Name":      "SETTINGS!$C$13",
    "Co_BankAcct":  "SETTINGS!$C$14",
    "Co_IBAN":      "SETTINGS!$C$15",
    "Co_Currency":  "SETTINGS!$C$16",
    "Co_BankName":  "SETTINGS!$C$17",
}
for nm, ref in named.items():
    dn = DefinedName(nm, attr_text=ref)
    wb.defined_names.add(dn)


# ═══════════════════════════════════════════════════════════════
# 3. BANK_RAW SHEET
# ═══════════════════════════════════════════════════════════════
ws_braw = wb.create_sheet("BANK_RAW")
ws_braw.sheet_properties.tabColor = P.EMERALD

bank_raw_cols = ["التاريخ\nDate", "الوصف\nDescription", "النوع\nType",
                 "الدائن\nCredit", "المدين\nDebit", "الرصيد\nBalance"]
bank_raw_widths = [15, 40, 18, 16, 16, 16]
set_col_widths(ws_braw, bank_raw_widths)

for i, h in enumerate(bank_raw_cols, 1):
    ws_braw.cell(row=1, column=i, value=h)
style_header_row(ws_braw, 1, len(bank_raw_cols), 36)

# Pre-format data area
for r in range(2, DATA_ROWS + 2):
    for c in range(1, 7):
        cell = ws_braw.cell(row=r, column=c)
        cell.border = thin_border
        cell.font = data_font
        if c == 1:
            cell.number_format = DATE_FMT
        elif c >= 4:
            cell.number_format = NUM_FMT

add_table(ws_braw, f"A1:F{DATA_ROWS+1}", "tbl_BankRaw", "TableStyleLight9")
ws_braw.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════════
# 4. GL_RAW SHEET
# ═══════════════════════════════════════════════════════════════
ws_graw = wb.create_sheet("GL_RAW")
ws_graw.sheet_properties.tabColor = P.AZURE2

gl_raw_cols = ["التاريخ\nDate", "رقم القيد\nJE Number", "مفتاح الترحيل\nPosting Key",
               "نوع القيد\nJE Type", "المبلغ\nAmount", "رقم الحساب\nGL Account",
               "اسم الحساب\nGL Name"]
gl_raw_widths = [15, 18, 16, 16, 16, 22, 30]
set_col_widths(ws_graw, gl_raw_widths)

for i, h in enumerate(gl_raw_cols, 1):
    ws_graw.cell(row=1, column=i, value=h)
style_header_row(ws_graw, 1, len(gl_raw_cols), 36)

for r in range(2, DATA_ROWS + 2):
    for c in range(1, 8):
        cell = ws_graw.cell(row=r, column=c)
        cell.border = thin_border
        cell.font = data_font
        if c == 1:
            cell.number_format = DATE_FMT
        elif c == 5:
            cell.number_format = NUM_FMT

add_table(ws_graw, f"A1:G{DATA_ROWS+1}", "tbl_GLRaw", "TableStyleLight9")
ws_graw.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════════
# 5. BANK_CLEAN SHEET
# ═══════════════════════════════════════════════════════════════
ws_bclean = wb.create_sheet("BANK_CLEAN")
ws_bclean.sheet_properties.tabColor = P.EMERALD

bc_cols = ["معرف\nRow ID", "التاريخ\nDate", "الوصف\nDescription", "النوع\nType",
           "الدائن\nCredit", "المدين\nDebit", "الرصيد\nBalance",
           "الجانب\nSide", "المبلغ\nAmount", "المبلغ المطلق\nAbs Amount",
           "نوع المطابقة\nMatch Key", "مستخدم\nUsed", "ملاحظات\nNotes"]
bc_widths = [8, 14, 38, 16, 15, 15, 15, 8, 15, 15, 12, 8, 20]
set_col_widths(ws_bclean, bc_widths)

for i, h in enumerate(bc_cols, 1):
    ws_bclean.cell(row=1, column=i, value=h)
style_header_row(ws_bclean, 1, len(bc_cols), 36)

# Formulas for BANK_CLEAN
for r in range(2, DATA_ROWS + 2):
    row = r
    br = r  # same row in BANK_RAW
    # A: Row ID
    ws_bclean.cell(row=row, column=1,
        value=f'=IF(BANK_RAW!A{br}="","",ROW()-1)')
    # B: Date (standardize)
    ws_bclean.cell(row=row, column=2,
        value=f'=IF(BANK_RAW!A{br}="","",IF(ISNUMBER(BANK_RAW!A{br}),BANK_RAW!A{br},DATEVALUE(TEXT(BANK_RAW!A{br},"YYYY-MM-DD"))))')
    ws_bclean.cell(row=row, column=2).number_format = DATE_FMT
    # C: Description
    ws_bclean.cell(row=row, column=3,
        value=f'=IF(BANK_RAW!A{br}="","",TRIM(BANK_RAW!B{br}))')
    # D: Type
    ws_bclean.cell(row=row, column=4,
        value=f'=IF(BANK_RAW!A{br}="","",TRIM(BANK_RAW!C{br}))')
    # E: Credit
    ws_bclean.cell(row=row, column=5,
        value=f'=IF(BANK_RAW!A{br}="","",IFERROR(ABS(BANK_RAW!D{br}),0))')
    ws_bclean.cell(row=row, column=5).number_format = NUM_FMT
    # F: Debit
    ws_bclean.cell(row=row, column=6,
        value=f'=IF(BANK_RAW!A{br}="","",IFERROR(ABS(BANK_RAW!E{br}),0))')
    ws_bclean.cell(row=row, column=6).number_format = NUM_FMT
    # G: Balance
    ws_bclean.cell(row=row, column=7,
        value=f'=IF(BANK_RAW!A{br}="","",IFERROR(BANK_RAW!F{br},0))')
    ws_bclean.cell(row=row, column=7).number_format = NUM_FMT
    # H: Side
    ws_bclean.cell(row=row, column=8,
        value=f'=IF(BANK_RAW!A{br}="","",IF(E{row}>0,"CR",IF(F{row}>0,"DR","CR")))')
    # I: Amount = Credit - Debit
    ws_bclean.cell(row=row, column=9,
        value=f'=IF(BANK_RAW!A{br}="","",E{row}-F{row})')
    ws_bclean.cell(row=row, column=9).number_format = NUM_FMT
    # J: Absolute Amount
    ws_bclean.cell(row=row, column=10,
        value=f'=IF(BANK_RAW!A{br}="","",ABS(I{row}))')
    ws_bclean.cell(row=row, column=10).number_format = NUM_FMT
    # K: Match Key Type (CR bank = IN matching, DR bank = OUT matching)
    ws_bclean.cell(row=row, column=11,
        value=f'=IF(BANK_RAW!A{br}="","",IF(H{row}="CR","IN","OUT"))')
    # L: Used flag (filled by MATCH_ENGINE reference)
    ws_bclean.cell(row=row, column=12, value="")
    # M: Notes
    ws_bclean.cell(row=row, column=13, value="")

    for c in range(1, 14):
        cell = ws_bclean.cell(row=row, column=c)
        cell.border = thin_border
        cell.font = data_font if c != 1 else mono_font

add_table(ws_bclean, f"A1:M{DATA_ROWS+1}", "tbl_BankClean", "TableStyleLight9")
ws_bclean.freeze_panes = "A2"


# ═══════════════════════════════════════════════════════════════
# 6. GL_CLEAN SHEET
# ═══════════════════════════════════════════════════════════════
ws_gclean = wb.create_sheet("GL_CLEAN")
ws_gclean.sheet_properties.tabColor = P.AZURE2

gc_cols = ["معرف\nRow ID", "التاريخ\nDate", "رقم القيد\nJE Number",
           "مفتاح الترحيل\nPosting Key", "نوع القيد\nJE Type",
           "رقم الحساب\nGL Account", "اسم الحساب\nGL Name",
           "المبلغ\nAmount", "المبلغ المطلق\nAbs Amount",
           "الاتجاه\nDirection", "ملاحظات\nNotes"]
gc_widths = [8, 14, 18, 14, 14, 22, 28, 15, 15, 10, 20]
set_col_widths(ws_gclean, gc_widths)

for i, h in enumerate(gc_cols, 1):
    ws_gclean.cell(row=1, column=i, value=h)
style_header_row(ws_gclean, 1, len(gc_cols), 36)

for r in range(2, DATA_ROWS + 2):
    row = r
    gr = r
    # A: Row ID
    ws_gclean.cell(row=row, column=1,
        value=f'=IF(GL_RAW!A{gr}="","",ROW()-1)')
    # B: Date
    ws_gclean.cell(row=row, column=2,
        value=f'=IF(GL_RAW!A{gr}="","",IF(ISNUMBER(GL_RAW!A{gr}),GL_RAW!A{gr},DATEVALUE(TEXT(GL_RAW!A{gr},"YYYY-MM-DD"))))')
    ws_gclean.cell(row=row, column=2).number_format = DATE_FMT
    # C: JE Number
    ws_gclean.cell(row=row, column=3,
        value=f'=IF(GL_RAW!A{gr}="","",TRIM(GL_RAW!B{gr}))')
    # D: Posting Key
    ws_gclean.cell(row=row, column=4,
        value=f'=IF(GL_RAW!A{gr}="","",TRIM(GL_RAW!C{gr}))')
    # E: JE Type
    ws_gclean.cell(row=row, column=5,
        value=f'=IF(GL_RAW!A{gr}="","",TRIM(GL_RAW!D{gr}))')
    # F: GL Account
    ws_gclean.cell(row=row, column=6,
        value=f'=IF(GL_RAW!A{gr}="","",TRIM(GL_RAW!F{gr}))')
    # G: GL Name
    ws_gclean.cell(row=row, column=7,
        value=f'=IF(GL_RAW!A{gr}="","",TRIM(GL_RAW!G{gr}))')
    # H: Amount
    ws_gclean.cell(row=row, column=8,
        value=f'=IF(GL_RAW!A{gr}="","",IFERROR(GL_RAW!E{gr},0))')
    ws_gclean.cell(row=row, column=8).number_format = NUM_FMT
    # I: Abs Amount
    ws_gclean.cell(row=row, column=9,
        value=f'=IF(GL_RAW!A{gr}="","",ABS(H{row}))')
    ws_gclean.cell(row=row, column=9).number_format = NUM_FMT
    # J: Direction (OUT if account contains "Out" or "11025002" or amount < 0)
    ws_gclean.cell(row=row, column=10,
        value=f'=IF(GL_RAW!A{gr}="","",IF(OR(ISNUMBER(SEARCH("Out",F{row})),ISNUMBER(SEARCH("11025002",F{row})),H{row}<0),"OUT","IN"))')
    # K: Notes
    ws_gclean.cell(row=row, column=11, value="")

    for c in range(1, 12):
        cell = ws_gclean.cell(row=row, column=c)
        cell.border = thin_border
        cell.font = data_font if c != 1 else mono_font

add_table(ws_gclean, f"A1:K{DATA_ROWS+1}", "tbl_GLClean", "TableStyleLight9")
ws_gclean.freeze_panes = "A2"


# ═══════════════════════════════════════════════════════════════
# 7. MATCH_ENGINE SHEET
# ═══════════════════════════════════════════════════════════════
ws_match = wb.create_sheet("MATCH_ENGINE")
ws_match.sheet_properties.tabColor = P.CRIMSON

me_cols = [
    "GL Row\nمعرف GL", "GL Date\nتاريخ GL", "GL JE\nرقم القيد",
    "GL Amount\nمبلغ GL", "GL Abs Amt\nالمبلغ المطلق", "GL Direction\nالاتجاه",
    "Bank Pool\nمجموعة البنك", "Best Bank Row\nأفضل صف بنك",
    "Best Bank Date\nتاريخ البنك", "Best Bank Amt\nمبلغ البنك",
    "Amt Diff\nفرق المبلغ", "Days Diff\nفرق الأيام",
    "Match Score\nنسبة التطابق", "Status\nالحالة"
]
me_widths = [10, 14, 16, 15, 15, 10, 10, 12, 14, 15, 13, 11, 13, 16]
set_col_widths(ws_match, me_widths)

for i, h in enumerate(me_cols, 1):
    ws_match.cell(row=1, column=i, value=h)
style_header_row(ws_match, 1, len(me_cols), 36)

# The matching engine uses helper formulas per GL row.
# For each GL row, we find the best matching bank row from BANK_CLEAN
# based on direction (IN->CR, OUT->DR), amount diff, and date diff.
# We use a scoring approach: Score = 100 - (pct_diff * 200) - (days * 3)

for r in range(2, DATA_ROWS + 2):
    row = r
    gc = r  # corresponding GL_CLEAN row

    # A: GL Row ID
    ws_match.cell(row=row, column=1,
        value=f'=IF(GL_CLEAN!A{gc}="","",GL_CLEAN!A{gc})')

    # B: GL Date
    ws_match.cell(row=row, column=2,
        value=f'=IF(A{row}="","",GL_CLEAN!B{gc})')
    ws_match.cell(row=row, column=2).number_format = DATE_FMT

    # C: GL JE
    ws_match.cell(row=row, column=3,
        value=f'=IF(A{row}="","",GL_CLEAN!C{gc})')

    # D: GL Amount
    ws_match.cell(row=row, column=4,
        value=f'=IF(A{row}="","",GL_CLEAN!H{gc})')
    ws_match.cell(row=row, column=4).number_format = NUM_FMT

    # E: GL Abs Amount
    ws_match.cell(row=row, column=5,
        value=f'=IF(A{row}="","",GL_CLEAN!I{gc})')
    ws_match.cell(row=row, column=5).number_format = NUM_FMT

    # F: GL Direction
    ws_match.cell(row=row, column=6,
        value=f'=IF(A{row}="","",GL_CLEAN!J{gc})')

    # G: Bank Pool (which side to match against)
    ws_match.cell(row=row, column=7,
        value=f'=IF(A{row}="","",IF(F{row}="OUT","DR","CR"))')

    # H: Best Bank Row - MINIFS approach to find closest amount match
    # Find the bank row with minimum absolute amount difference in the correct pool
    ws_match.cell(row=row, column=8,
        value=f'=IF(A{row}="","",IFERROR(INDEX(BANK_CLEAN!$A$2:$A${DATA_ROWS+1},MATCH(MINIFS(BANK_CLEAN!$J$2:$J${DATA_ROWS+1},BANK_CLEAN!$H$2:$H${DATA_ROWS+1},G{row},BANK_CLEAN!$L$2:$L${DATA_ROWS+1},""),BANK_CLEAN!$J$2:$J${DATA_ROWS+1},0)),""))')

    # I: Best Bank Date
    ws_match.cell(row=row, column=9,
        value=f'=IF(H{row}="","",INDEX(BANK_CLEAN!$B$2:$B${DATA_ROWS+1},H{row}))')
    ws_match.cell(row=row, column=9).number_format = DATE_FMT

    # J: Best Bank Amount (credit or debit based on direction)
    ws_match.cell(row=row, column=10,
        value=f'=IF(H{row}="","",IF(F{row}="OUT",INDEX(BANK_CLEAN!$F$2:$F${DATA_ROWS+1},H{row}),INDEX(BANK_CLEAN!$E$2:$E${DATA_ROWS+1},H{row})))')
    ws_match.cell(row=row, column=10).number_format = NUM_FMT

    # K: Amount Difference
    ws_match.cell(row=row, column=11,
        value=f'=IF(H{row}="","",ABS(E{row}-J{row}))')
    ws_match.cell(row=row, column=11).number_format = NUM_FMT

    # L: Days Difference
    ws_match.cell(row=row, column=12,
        value=f'=IF(OR(H{row}="",B{row}="",I{row}=""),"",ABS(B{row}-I{row}))')

    # M: Match Score = 100 - (pct_diff * 200) - (days * 3)
    ws_match.cell(row=row, column=13,
        value=f'=IF(H{row}="",0,LET(pct,IF(E{row}>0,K{row}/E{row},1),days,IFERROR(L{row},999),100-pct*200-days*3))')
    ws_match.cell(row=row, column=13).number_format = '0.0'

    # N: Status
    ws_match.cell(row=row, column=14,
        value=f'=IF(A{row}="","",IF(H{row}="","Unmatched GL",IF(AND(K{row}<=Tol_Amount,L{row}<=Tol_Days),"Matched",IF(M{row}>Tol_NearPct,"Near Match","Unmatched GL"))))')

    for c in range(1, 15):
        cell = ws_match.cell(row=row, column=c)
        cell.border = thin_border
        cell.font = mono_font if c in (1, 4, 5, 10, 11, 13) else data_font

# Conditional formatting for status column (N)
from openpyxl.formatting.rule import CellIsRule
green_fill = PatternFill(start_color=P.EMERALD3, end_color=P.EMERALD3, fill_type='solid')
amber_fill = PatternFill(start_color=P.AMBER3, end_color=P.AMBER3, fill_type='solid')
red_fill = PatternFill(start_color=P.CRIMSON3, end_color=P.CRIMSON3, fill_type='solid')
blue_fill = PatternFill(start_color=P.AZURE4, end_color=P.AZURE4, fill_type='solid')

green_font = Font(name='Calibri', size=10, bold=True, color=P.EMERALD)
amber_font = Font(name='Calibri', size=10, bold=True, color=P.AMBER)
red_font = Font(name='Calibri', size=10, bold=True, color=P.CRIMSON)
blue_font = Font(name='Calibri', size=10, bold=True, color=P.AZURE)

status_range = f"N2:N{DATA_ROWS+1}"
ws_match.conditional_formatting.add(status_range,
    CellIsRule(operator='equal', formula=['"Matched"'], fill=green_fill, font=green_font))
ws_match.conditional_formatting.add(status_range,
    CellIsRule(operator='equal', formula=['"Near Match"'], fill=amber_fill, font=amber_font))
ws_match.conditional_formatting.add(status_range,
    CellIsRule(operator='equal', formula=['"Unmatched GL"'], fill=red_fill, font=red_font))

add_table(ws_match, f"A1:N{DATA_ROWS+1}", "tbl_MatchEngine", "TableStyleLight9")
ws_match.freeze_panes = "A2"


# ═══════════════════════════════════════════════════════════════
# 8. RECON_RESULTS SHEET
# ═══════════════════════════════════════════════════════════════
ws_recon = wb.create_sheet("RECON_RESULTS")
ws_recon.sheet_properties.tabColor = P.AZURE

rr_cols = [
    "الحالة\nStatus", "رقم القيد\nGL JE", "تاريخ GL\nGL Date",
    "مبلغ GL\nGL Amount", "الاتجاه\nDirection",
    "تاريخ البنك\nBank Date", "الدائن\nBank Credit", "المدين\nBank Debit",
    "مبلغ البنك\nBank Amount", "الفرق\nDifference", "الأيام\nDays Apart",
    "نسبة التطابق\nScore", "الحساب\nGL Account", "اسم الحساب\nGL Name",
    "وصف البنك\nBank Desc", "ملاحظات المراجع\nReviewer Notes"
]
rr_widths = [16, 16, 14, 15, 10, 14, 15, 15, 15, 13, 10, 10, 20, 25, 30, 20]
set_col_widths(ws_recon, rr_widths)

for i, h in enumerate(rr_cols, 1):
    ws_recon.cell(row=1, column=i, value=h)
style_header_row(ws_recon, 1, len(rr_cols), 36)

for r in range(2, DATA_ROWS + 2):
    row = r
    me = r  # MATCH_ENGINE row
    gc = r  # GL_CLEAN row

    # A: Status
    ws_recon.cell(row=row, column=1,
        value=f'=IF(MATCH_ENGINE!A{me}="","",MATCH_ENGINE!N{me})')
    # B: GL JE
    ws_recon.cell(row=row, column=2,
        value=f'=IF(MATCH_ENGINE!A{me}="","",MATCH_ENGINE!C{me})')
    # C: GL Date
    ws_recon.cell(row=row, column=3,
        value=f'=IF(MATCH_ENGINE!A{me}="","",MATCH_ENGINE!B{me})')
    ws_recon.cell(row=row, column=3).number_format = DATE_FMT
    # D: GL Amount
    ws_recon.cell(row=row, column=4,
        value=f'=IF(MATCH_ENGINE!A{me}="","",MATCH_ENGINE!D{me})')
    ws_recon.cell(row=row, column=4).number_format = SAR_FMT
    # E: Direction
    ws_recon.cell(row=row, column=5,
        value=f'=IF(MATCH_ENGINE!A{me}="","",MATCH_ENGINE!F{me})')
    # F: Bank Date
    ws_recon.cell(row=row, column=6,
        value=f'=IF(MATCH_ENGINE!A{me}="","",MATCH_ENGINE!I{me})')
    ws_recon.cell(row=row, column=6).number_format = DATE_FMT
    # G: Bank Credit
    ws_recon.cell(row=row, column=7,
        value=f'=IF(MATCH_ENGINE!H{me}="","",IF(MATCH_ENGINE!F{me}="IN",MATCH_ENGINE!J{me},""))')
    ws_recon.cell(row=row, column=7).number_format = SAR_FMT
    # H: Bank Debit
    ws_recon.cell(row=row, column=8,
        value=f'=IF(MATCH_ENGINE!H{me}="","",IF(MATCH_ENGINE!F{me}="OUT",MATCH_ENGINE!J{me},""))')
    ws_recon.cell(row=row, column=8).number_format = SAR_FMT
    # I: Bank Amount (signed)
    ws_recon.cell(row=row, column=9,
        value=f'=IF(MATCH_ENGINE!H{me}="","",MATCH_ENGINE!J{me})')
    ws_recon.cell(row=row, column=9).number_format = SAR_FMT
    # J: Difference
    ws_recon.cell(row=row, column=10,
        value=f'=IF(MATCH_ENGINE!A{me}="","",MATCH_ENGINE!K{me})')
    ws_recon.cell(row=row, column=10).number_format = NUM_FMT
    # K: Days Apart
    ws_recon.cell(row=row, column=11,
        value=f'=IF(MATCH_ENGINE!A{me}="","",MATCH_ENGINE!L{me})')
    # L: Score
    ws_recon.cell(row=row, column=12,
        value=f'=IF(MATCH_ENGINE!A{me}="","",MATCH_ENGINE!M{me})')
    ws_recon.cell(row=row, column=12).number_format = '0.0'
    # M: GL Account
    ws_recon.cell(row=row, column=13,
        value=f'=IF(MATCH_ENGINE!A{me}="","",GL_CLEAN!F{gc})')
    # N: GL Name
    ws_recon.cell(row=row, column=14,
        value=f'=IF(MATCH_ENGINE!A{me}="","",GL_CLEAN!G{gc})')
    # O: Bank Description
    ws_recon.cell(row=row, column=15,
        value=f'=IF(MATCH_ENGINE!H{me}="","",INDEX(BANK_CLEAN!$C$2:$C${DATA_ROWS+1},MATCH_ENGINE!H{me}))')
    # P: Reviewer Notes (manual entry)
    ws_recon.cell(row=row, column=16, value="")

    for c in range(1, 17):
        cell = ws_recon.cell(row=row, column=c)
        cell.border = thin_border
        cell.font = data_font

# Conditional formatting for status
sr = f"A2:A{DATA_ROWS+1}"
ws_recon.conditional_formatting.add(sr,
    CellIsRule(operator='equal', formula=['"Matched"'], fill=green_fill, font=green_font))
ws_recon.conditional_formatting.add(sr,
    CellIsRule(operator='equal', formula=['"Near Match"'], fill=amber_fill, font=amber_font))
ws_recon.conditional_formatting.add(sr,
    CellIsRule(operator='equal', formula=['"Unmatched GL"'], fill=red_fill, font=red_font))
ws_recon.conditional_formatting.add(sr,
    CellIsRule(operator='equal', formula=['"Unmatched Bank"'], fill=blue_fill, font=blue_font))

# Full row conditional formatting
for status, fill_s, font_s in [
    ("Matched", green_fill, green_font),
    ("Near Match", amber_fill, amber_font),
    ("Unmatched GL", red_fill, red_font),
]:
    ws_recon.conditional_formatting.add(
        f"A2:P{DATA_ROWS+1}",
        FormulaRule(formula=[f'$A2="{status}"'], fill=fill_s)
    )

add_table(ws_recon, f"A1:P{DATA_ROWS+1}", "tbl_ReconResults", "TableStyleLight9")
ws_recon.freeze_panes = "A2"


# ═══════════════════════════════════════════════════════════════
# 9. DASHBOARD SHEET
# ═══════════════════════════════════════════════════════════════
ws_dash = wb.create_sheet("DASHBOARD")
ws_dash.sheet_properties.tabColor = P.AZURE
ws_dash.sheet_view.showGridLines = False
set_col_widths(ws_dash, [2, 22, 22, 22, 22, 22, 22, 2])

# Background fill for entire visible area
dash_bg = PatternFill(start_color=P.CREAM, end_color=P.CREAM, fill_type='solid')
white_fill = PatternFill(start_color=P.WHITE, end_color=P.WHITE, fill_type='solid')
azure_fill_d = PatternFill(start_color=P.AZURE, end_color=P.AZURE, fill_type='solid')
ink_fill = PatternFill(start_color=P.INK, end_color=P.INK, fill_type='solid')

for r in range(1, 65):
    for c in range(1, 9):
        ws_dash.cell(row=r, column=c).fill = dash_bg

# ─── HEADER ───
for c in range(1, 9):
    ws_dash.cell(row=2, column=c).fill = ink_fill
    ws_dash.cell(row=3, column=c).fill = ink_fill

ws_dash.merge_cells('B2:D3')
h1 = ws_dash['B2']
h1.value = '=Co_Name'
h1.font = Font(name='Calibri', size=16, bold=True, color=P.WHITE)
h1.alignment = Alignment(horizontal='right', vertical='center')

ws_dash.merge_cells('E2:G2')
h2 = ws_dash['E2']
h2.value = "تقرير مطابقة الحسابات البنكية"
h2.font = Font(name='Calibri', size=12, color="FFFFFF")
h2.alignment = Alignment(horizontal='left', vertical='center')

ws_dash.merge_cells('E3:G3')
h3 = ws_dash['E3']
h3.value = '=Co_BankName&" — "&Co_IBAN&" — "&Rpt_Period'
h3.font = Font(name='Consolas', size=9, color=P.LIGHT)
h3.alignment = Alignment(horizontal='left', vertical='center')

# Azure accent line
for c in range(1, 9):
    cell = ws_dash.cell(row=4, column=c)
    cell.fill = PatternFill(start_color=P.AZURE2, end_color=P.AZURE2, fill_type='solid')
ws_dash.row_dimensions[4].height = 4

# ─── KPI SECTION ───
kpi_start = 6
ws_dash.merge_cells('B6:G6')
ws_dash['B6'].value = "مؤشرات الأداء الرئيسية — KEY PERFORMANCE INDICATORS"
ws_dash['B6'].font = Font(name='Consolas', size=9, color=P.MID, bold=False)
ws_dash['B6'].alignment = Alignment(horizontal='right')

kpis = [
    # (col, label_ar, label_en, formula, color, fmt)
    (2, "معدل المطابقة", "MATCH RATE",
     f'=IFERROR((COUNTIF(MATCH_ENGINE!N$2:N${DATA_ROWS+1},"Matched")+COUNTIF(MATCH_ENGINE!N$2:N${DATA_ROWS+1},"Near Match"))/COUNTIF(MATCH_ENGINE!N$2:N${DATA_ROWS+1},"<>"),0)',
     P.EMERALD, PCT_FMT),
    (3, "مطابق", "MATCHED",
     f'=COUNTIF(MATCH_ENGINE!N$2:N${DATA_ROWS+1},"Matched")', P.EMERALD, '#,##0'),
    (4, "قريب", "NEAR MATCH",
     f'=COUNTIF(MATCH_ENGINE!N$2:N${DATA_ROWS+1},"Near Match")', P.AMBER, '#,##0'),
    (5, "GL غير مطابق", "UNMATCHED GL",
     f'=COUNTIF(MATCH_ENGINE!N$2:N${DATA_ROWS+1},"Unmatched GL")', P.CRIMSON, '#,##0'),
    (6, "بنك غير مطابق", "UNMATCHED BANK",
     f'=COUNTIFS(BANK_CLEAN!A$2:A${DATA_ROWS+1},"<>",BANK_CLEAN!L$2:L${DATA_ROWS+1},"")-COUNTIF(BANK_CLEAN!A$2:A${DATA_ROWS+1},"")', P.AZURE, '#,##0'),
    (7, "إجمالي الاستثناءات", "EXCEPTIONS",
     f'=COUNTIF(MATCH_ENGINE!N$2:N${DATA_ROWS+1},"Unmatched GL")+COUNTIF(MATCH_ENGINE!N$2:N${DATA_ROWS+1},"Near Match")', P.CRIMSON, '#,##0'),
]

for col, ar, en, formula, color, nfmt in kpis:
    r_label = kpi_start + 1
    r_en = kpi_start + 2
    r_val = kpi_start + 3

    cell_ar = ws_dash.cell(row=r_label, column=col, value=ar)
    cell_ar.font = Font(name='Calibri', size=10, color=P.MID)
    cell_ar.alignment = Alignment(horizontal='center')
    cell_ar.fill = white_fill

    cell_en = ws_dash.cell(row=r_en, column=col, value=en)
    cell_en.font = Font(name='Consolas', size=8, color=P.LIGHT)
    cell_en.alignment = Alignment(horizontal='center')
    cell_en.fill = white_fill

    cell_val = ws_dash.cell(row=r_val, column=col, value=formula)
    cell_val.font = Font(name='Consolas', size=20, bold=True, color=color)
    cell_val.number_format = nfmt
    cell_val.alignment = Alignment(horizontal='center', vertical='center')
    cell_val.fill = white_fill

    # Border the KPI card
    for rr in range(r_label, r_val + 1):
        for cc in [col]:
            ws_dash.cell(row=rr, column=cc).border = Border(
                left=Side(style='thin', color=P.RULE),
                right=Side(style='thin', color=P.RULE),
                top=Side(style='thin', color=P.RULE) if rr == r_label else Side(style=None),
                bottom=Side(style='thin', color=P.RULE) if rr == r_val else Side(style=None),
            )

ws_dash.row_dimensions[kpi_start + 3].height = 40

# ─── RECONCILIATION WATERFALL ───
wf_start = 12
ws_dash.merge_cells(f'B{wf_start}:D{wf_start}')
ws_dash[f'B{wf_start}'].value = "بيان التوفيق — RECONCILIATION WATERFALL"
ws_dash[f'B{wf_start}'].font = Font(name='Consolas', size=9, color=P.MID)
ws_dash[f'B{wf_start}'].alignment = Alignment(horizontal='right')

wf_items = [
    ("الدائن البنكي", "Bank Credit",
     f'=SUMPRODUCT((BANK_CLEAN!A$2:A${DATA_ROWS+1}<>"")*BANK_CLEAN!E$2:E${DATA_ROWS+1})',
     P.EMERALD, False, False),
    ("ناقص: المدين البنكي", "Less: Bank Debit",
     f'=SUMPRODUCT((BANK_CLEAN!A$2:A${DATA_ROWS+1}<>"")*BANK_CLEAN!F$2:F${DATA_ROWS+1})',
     P.CRIMSON, True, False),
    ("صافي حركة البنك", "Net Bank Movement",
     f'=B{wf_start+2}-B{wf_start+3}',
     P.INK, False, False),
    ("GL الوارد", "GL Inward",
     f'=SUMPRODUCT((GL_CLEAN!J$2:J${DATA_ROWS+1}="IN")*(GL_CLEAN!A$2:A${DATA_ROWS+1}<>"")*GL_CLEAN!H$2:H${DATA_ROWS+1})',
     P.EMERALD, False, False),
    ("ناقص: GL الصادر", "GL Outward",
     f'=SUMPRODUCT((GL_CLEAN!J$2:J${DATA_ROWS+1}="OUT")*(GL_CLEAN!A$2:A${DATA_ROWS+1}<>"")*ABS(GL_CLEAN!H$2:H${DATA_ROWS+1}))',
     P.CRIMSON, True, False),
    ("صافي حركة GL", "Net GL Movement",
     f'=B{wf_start+5}-B{wf_start+6}',
     P.INK, False, False),
    ("الفرق (بنك - GL)", "Difference",
     f'=ABS(B{wf_start+4}-B{wf_start+7})',
     P.CRIMSON, False, True),
]

for idx, (ar, en, formula, color, indent, is_final) in enumerate(wf_items):
    r = wf_start + 1 + idx
    ws_dash.row_dimensions[r].height = 28

    label_cell = ws_dash.cell(row=r, column=3, value=ar)
    label_cell.font = Font(name='Calibri', size=11,
                           color=P.SLATE, bold=is_final)
    label_cell.alignment = Alignment(horizontal='right', indent=2 if indent else 0)
    label_cell.fill = white_fill

    en_cell = ws_dash.cell(row=r, column=4, value=en)
    en_cell.font = Font(name='Consolas', size=9, color=P.LIGHT)
    en_cell.alignment = Alignment(horizontal='right')
    en_cell.fill = white_fill

    val_cell = ws_dash.cell(row=r, column=2, value=formula)
    val_cell.font = Font(name='Consolas', size=12, bold=True if is_final else False, color=color)
    val_cell.number_format = SAR_FMT
    val_cell.alignment = Alignment(horizontal='left')
    val_cell.fill = white_fill

    for cc in range(2, 5):
        ws_dash.cell(row=r, column=cc).border = Border(
            bottom=Side(style='double' if is_final else 'thin', color=P.INK if is_final else P.PALE),
            top=Side(style='double' if is_final else None, color=P.INK if is_final else None),
        )

# ─── STATUS SUMMARY ───
ss_start = wf_start + 10
ws_dash.merge_cells(f'B{ss_start}:D{ss_start}')
ws_dash[f'B{ss_start}'].value = "ملخص الحالات — STATUS SUMMARY"
ws_dash[f'B{ss_start}'].font = Font(name='Consolas', size=9, color=P.MID)
ws_dash[f'B{ss_start}'].alignment = Alignment(horizontal='right')

status_cards = [
    ("Matched", "مطابق", P.EMERALD, P.EMERALD3, f'=COUNTIF(MATCH_ENGINE!N$2:N${DATA_ROWS+1},"Matched")'),
    ("Near Match", "قريب", P.AMBER, P.AMBER3, f'=COUNTIF(MATCH_ENGINE!N$2:N${DATA_ROWS+1},"Near Match")'),
    ("Unmatched GL", "GL غير مطابق", P.CRIMSON, P.CRIMSON3, f'=COUNTIF(MATCH_ENGINE!N$2:N${DATA_ROWS+1},"Unmatched GL")'),
    ("Unmatched Bank", "بنك غير مطابق", P.AZURE, P.AZURE4,
     f'=COUNTIFS(BANK_CLEAN!A$2:A${DATA_ROWS+1},"<>",BANK_CLEAN!L$2:L${DATA_ROWS+1},"")-COUNTIF(BANK_CLEAN!A$2:A${DATA_ROWS+1},"")'),
]

for idx, (en, ar, color, bg_color, formula) in enumerate(status_cards):
    col = 2 + idx
    r1 = ss_start + 1
    r2 = ss_start + 2
    r3 = ss_start + 3

    card_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

    val_c = ws_dash.cell(row=r1, column=col, value=formula)
    val_c.font = Font(name='Consolas', size=26, bold=True, color=color)
    val_c.alignment = Alignment(horizontal='center', vertical='center')
    val_c.fill = card_fill

    ar_c = ws_dash.cell(row=r2, column=col, value=ar)
    ar_c.font = Font(name='Calibri', size=10, color=P.MID)
    ar_c.alignment = Alignment(horizontal='center')
    ar_c.fill = card_fill

    en_c = ws_dash.cell(row=r3, column=col, value=en.upper())
    en_c.font = Font(name='Consolas', size=8, color=color)
    en_c.alignment = Alignment(horizontal='center')
    en_c.fill = card_fill

    for rr in range(r1, r3 + 1):
        ws_dash.cell(row=rr, column=col).border = Border(
            left=Side(style='thin', color=P.RULE),
            right=Side(style='thin', color=P.RULE),
            top=Side(style='thin', color=P.RULE) if rr == r1 else Side(style=None),
            bottom=Side(style='thin', color=P.RULE) if rr == r3 else Side(style=None),
        )

ws_dash.row_dimensions[ss_start + 1].height = 45

# ─── EXCEPTION EXPOSURE ───
ex_start = ss_start + 6
ws_dash.merge_cells(f'B{ex_start}:D{ex_start}')
ws_dash[f'B{ex_start}'].value = "التعرض للاستثناءات — EXCEPTION EXPOSURE"
ws_dash[f'B{ex_start}'].font = Font(name='Consolas', size=9, color=P.MID)
ws_dash[f'B{ex_start}'].alignment = Alignment(horizontal='right')

exp_items = [
    ("إجمالي البنك غير المطابق", "Unmatched Bank Value",
     f'=SUMPRODUCT((RECON_RESULTS!A$2:A${DATA_ROWS+1}="Unmatched Bank")*(RECON_RESULTS!I$2:I${DATA_ROWS+1}))',
     P.CRIMSON),
    ("إجمالي GL غير المطابق", "Unmatched GL Value",
     f'=SUMPRODUCT((RECON_RESULTS!A$2:A${DATA_ROWS+1}="Unmatched GL")*ABS(RECON_RESULTS!D$2:D${DATA_ROWS+1}))',
     P.CRIMSON),
    ("إجمالي المطابقات القريبة", "Near Match Exposure",
     f'=SUMPRODUCT((RECON_RESULTS!A$2:A${DATA_ROWS+1}="Near Match")*RECON_RESULTS!J$2:J${DATA_ROWS+1})',
     P.AMBER),
]

for idx, (ar, en, formula, color) in enumerate(exp_items):
    r = ex_start + 1 + idx
    ws_dash.row_dimensions[r].height = 28

    ws_dash.cell(row=r, column=3, value=ar).font = Font(name='Calibri', size=11, color=P.SLATE)
    ws_dash.cell(row=r, column=3).alignment = Alignment(horizontal='right')
    ws_dash.cell(row=r, column=3).fill = white_fill

    ws_dash.cell(row=r, column=4, value=en).font = Font(name='Consolas', size=9, color=P.LIGHT)
    ws_dash.cell(row=r, column=4).alignment = Alignment(horizontal='right')
    ws_dash.cell(row=r, column=4).fill = white_fill

    val_c = ws_dash.cell(row=r, column=2, value=formula)
    val_c.font = Font(name='Consolas', size=12, bold=True, color=color)
    val_c.number_format = SAR_FMT
    val_c.alignment = Alignment(horizontal='left')
    val_c.fill = white_fill

    for cc in range(2, 5):
        ws_dash.cell(row=r, column=cc).border = Border(
            bottom=Side(style='thin', color=P.PALE))

ws_dash.freeze_panes = "A5"


# ═══════════════════════════════════════════════════════════════
# 10. EXCEPTIONS SHEET
# ═══════════════════════════════════════════════════════════════
ws_exc = wb.create_sheet("EXCEPTIONS")
ws_exc.sheet_properties.tabColor = P.CRIMSON
set_col_widths(ws_exc, [2, 16, 14, 16, 15, 14, 15, 13, 10, 10, 12, 20, 10])

exc_section_fill = PatternFill(start_color=P.INK, end_color=P.INK, fill_type='solid')

# ─── Section 1: Unmatched Bank ───
r = 1
ws_exc.merge_cells(f'B{r}:L{r}')
ws_exc.cell(row=r, column=2, value="حركات بنكية بدون قيد GL — UNMATCHED BANK")
ws_exc.cell(row=r, column=2).font = Font(name='Calibri', size=12, bold=True, color=P.WHITE)
ws_exc.cell(row=r, column=2).alignment = Alignment(horizontal='right', vertical='center')
for c in range(2, 13):
    ws_exc.cell(row=r, column=c).fill = exc_section_fill
ws_exc.row_dimensions[r].height = 32

r = 2
exc_bank_headers = ["الحالة\nStatus", "التاريخ\nDate", "الوصف\nDescription",
                    "الدائن\nCredit", "المدين\nDebit", "الرصيد\nBalance",
                    "الجانب\nSide", "النوع\nType", "حالة المراجعة\nReview",
                    "الأولوية\nPriority", "ملاحظات\nComments"]
for i, h in enumerate(exc_bank_headers, 2):
    ws_exc.cell(row=r, column=i, value=h)
style_header_row(ws_exc, r, 12, 32)

EXC_ROWS = 200
for rr in range(3, 3 + EXC_ROWS):
    idx = rr - 2
    # These will show bank rows where BANK_CLEAN.Used is empty
    # Using a simplified approach: reference RECON_RESULTS for Unmatched Bank status
    ws_exc.cell(row=rr, column=2, value=f'=IFERROR(INDEX(RECON_RESULTS!$A$2:$A${DATA_ROWS+1},SMALL(IF(RECON_RESULTS!$A$2:$A${DATA_ROWS+1}="Unmatched Bank",ROW(RECON_RESULTS!$A$2:$A${DATA_ROWS+1})-1),{idx})),"")')
    ws_exc.cell(row=rr, column=3, value=f'=IF(B{rr}="","",INDEX(RECON_RESULTS!$F$2:$F${DATA_ROWS+1},SMALL(IF(RECON_RESULTS!$A$2:$A${DATA_ROWS+1}="Unmatched Bank",ROW(RECON_RESULTS!$A$2:$A${DATA_ROWS+1})-1),{idx})))')
    ws_exc.cell(row=rr, column=3).number_format = DATE_FMT
    ws_exc.cell(row=rr, column=4, value=f'=IF(B{rr}="","",INDEX(RECON_RESULTS!$O$2:$O${DATA_ROWS+1},SMALL(IF(RECON_RESULTS!$A$2:$A${DATA_ROWS+1}="Unmatched Bank",ROW(RECON_RESULTS!$A$2:$A${DATA_ROWS+1})-1),{idx})))')
    ws_exc.cell(row=rr, column=5, value=f'=IF(B{rr}="","",INDEX(RECON_RESULTS!$G$2:$G${DATA_ROWS+1},SMALL(IF(RECON_RESULTS!$A$2:$A${DATA_ROWS+1}="Unmatched Bank",ROW(RECON_RESULTS!$A$2:$A${DATA_ROWS+1})-1),{idx})))')
    ws_exc.cell(row=rr, column=5).number_format = SAR_FMT
    ws_exc.cell(row=rr, column=6, value=f'=IF(B{rr}="","",INDEX(RECON_RESULTS!$H$2:$H${DATA_ROWS+1},SMALL(IF(RECON_RESULTS!$A$2:$A${DATA_ROWS+1}="Unmatched Bank",ROW(RECON_RESULTS!$A$2:$A${DATA_ROWS+1})-1),{idx})))')
    ws_exc.cell(row=rr, column=6).number_format = SAR_FMT
    ws_exc.cell(row=rr, column=7, value="")
    ws_exc.cell(row=rr, column=8, value="")
    ws_exc.cell(row=rr, column=9, value="")
    # Review status dropdown
    ws_exc.cell(row=rr, column=10, value="")
    ws_exc.cell(row=rr, column=11, value="")
    ws_exc.cell(row=rr, column=12, value="")

    for c in range(2, 13):
        ws_exc.cell(row=rr, column=c).border = thin_border
        ws_exc.cell(row=rr, column=c).font = data_font

# Review status validation
dv_review = DataValidation(type="list", formula1='"Pending,Reviewed,Resolved,Escalated"',
                           allow_blank=True)
dv_review.error = "اختر حالة المراجعة"
dv_review.prompt = "اختر حالة المراجعة"
ws_exc.add_data_validation(dv_review)
dv_review.add(f'J3:J{2+EXC_ROWS}')

dv_priority = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
ws_exc.add_data_validation(dv_priority)
dv_priority.add(f'K3:K{2+EXC_ROWS}')

# ─── Section 2: Unmatched GL ───
gl_sec_start = 3 + EXC_ROWS + 2
r = gl_sec_start
ws_exc.merge_cells(f'B{r}:L{r}')
ws_exc.cell(row=r, column=2, value="قيود GL بدون حركة بنكية — UNMATCHED GL")
ws_exc.cell(row=r, column=2).font = Font(name='Calibri', size=12, bold=True, color=P.WHITE)
ws_exc.cell(row=r, column=2).alignment = Alignment(horizontal='right', vertical='center')
for c in range(2, 13):
    ws_exc.cell(row=r, column=c).fill = exc_section_fill
ws_exc.row_dimensions[r].height = 32

r = gl_sec_start + 1
exc_gl_headers = ["الحالة\nStatus", "تاريخ GL\nGL Date", "رقم القيد\nJE Number",
                  "مبلغ GL\nGL Amount", "الاتجاه\nDirection", "الحساب\nGL Account",
                  "اسم الحساب\nGL Name", "نوع القيد\nJE Type", "حالة المراجعة\nReview",
                  "الأولوية\nPriority", "ملاحظات\nComments"]
for i, h in enumerate(exc_gl_headers, 2):
    ws_exc.cell(row=r, column=i, value=h)
style_header_row(ws_exc, r, 12, 32)

for rr in range(gl_sec_start + 2, gl_sec_start + 2 + EXC_ROWS):
    idx = rr - gl_sec_start - 1
    ws_exc.cell(row=rr, column=2, value=f'=IFERROR(INDEX(RECON_RESULTS!$A$2:$A${DATA_ROWS+1},SMALL(IF(RECON_RESULTS!$A$2:$A${DATA_ROWS+1}="Unmatched GL",ROW(RECON_RESULTS!$A$2:$A${DATA_ROWS+1})-1),{idx})),"")')
    ws_exc.cell(row=rr, column=3, value=f'=IF(B{rr}="","",INDEX(RECON_RESULTS!$C$2:$C${DATA_ROWS+1},SMALL(IF(RECON_RESULTS!$A$2:$A${DATA_ROWS+1}="Unmatched GL",ROW(RECON_RESULTS!$A$2:$A${DATA_ROWS+1})-1),{idx})))')
    ws_exc.cell(row=rr, column=3).number_format = DATE_FMT
    ws_exc.cell(row=rr, column=4, value=f'=IF(B{rr}="","",INDEX(RECON_RESULTS!$B$2:$B${DATA_ROWS+1},SMALL(IF(RECON_RESULTS!$A$2:$A${DATA_ROWS+1}="Unmatched GL",ROW(RECON_RESULTS!$A$2:$A${DATA_ROWS+1})-1),{idx})))')
    ws_exc.cell(row=rr, column=5, value=f'=IF(B{rr}="","",INDEX(RECON_RESULTS!$D$2:$D${DATA_ROWS+1},SMALL(IF(RECON_RESULTS!$A$2:$A${DATA_ROWS+1}="Unmatched GL",ROW(RECON_RESULTS!$A$2:$A${DATA_ROWS+1})-1),{idx})))')
    ws_exc.cell(row=rr, column=5).number_format = SAR_FMT
    ws_exc.cell(row=rr, column=6, value=f'=IF(B{rr}="","",INDEX(RECON_RESULTS!$E$2:$E${DATA_ROWS+1},SMALL(IF(RECON_RESULTS!$A$2:$A${DATA_ROWS+1}="Unmatched GL",ROW(RECON_RESULTS!$A$2:$A${DATA_ROWS+1})-1),{idx})))')
    ws_exc.cell(row=rr, column=7, value=f'=IF(B{rr}="","",INDEX(RECON_RESULTS!$M$2:$M${DATA_ROWS+1},SMALL(IF(RECON_RESULTS!$A$2:$A${DATA_ROWS+1}="Unmatched GL",ROW(RECON_RESULTS!$A$2:$A${DATA_ROWS+1})-1),{idx})))')
    ws_exc.cell(row=rr, column=8, value=f'=IF(B{rr}="","",INDEX(RECON_RESULTS!$N$2:$N${DATA_ROWS+1},SMALL(IF(RECON_RESULTS!$A$2:$A${DATA_ROWS+1}="Unmatched GL",ROW(RECON_RESULTS!$A$2:$A${DATA_ROWS+1})-1),{idx})))')
    ws_exc.cell(row=rr, column=9, value="")
    ws_exc.cell(row=rr, column=10, value="")
    ws_exc.cell(row=rr, column=11, value="")
    ws_exc.cell(row=rr, column=12, value="")

    for c in range(2, 13):
        ws_exc.cell(row=rr, column=c).border = thin_border
        ws_exc.cell(row=rr, column=c).font = data_font

# ─── Section 3: Near Matches ───
nm_sec_start = gl_sec_start + 2 + EXC_ROWS + 2
r = nm_sec_start
ws_exc.merge_cells(f'B{r}:L{r}')
ws_exc.cell(row=r, column=2, value="مطابقات قريبة تحتاج مراجعة — NEAR MATCHES FOR REVIEW")
ws_exc.cell(row=r, column=2).font = Font(name='Calibri', size=12, bold=True, color=P.WHITE)
ws_exc.cell(row=r, column=2).alignment = Alignment(horizontal='right', vertical='center')
for c in range(2, 13):
    ws_exc.cell(row=r, column=c).fill = exc_section_fill
ws_exc.row_dimensions[r].height = 32

r = nm_sec_start + 1
exc_nm_headers = ["الحالة\nStatus", "رقم القيد\nGL JE", "تاريخ GL\nGL Date",
                  "مبلغ GL\nGL Amount", "تاريخ البنك\nBank Date", "مبلغ البنك\nBank Amount",
                  "الفرق\nDifference", "الأيام\nDays", "النسبة\nScore",
                  "حالة المراجعة\nReview", "ملاحظات\nComments"]
for i, h in enumerate(exc_nm_headers, 2):
    ws_exc.cell(row=r, column=i, value=h)
style_header_row(ws_exc, r, 12, 32)

for rr in range(nm_sec_start + 2, nm_sec_start + 2 + EXC_ROWS):
    idx = rr - nm_sec_start - 1
    ws_exc.cell(row=rr, column=2, value=f'=IFERROR(INDEX(RECON_RESULTS!$A$2:$A${DATA_ROWS+1},SMALL(IF(RECON_RESULTS!$A$2:$A${DATA_ROWS+1}="Near Match",ROW(RECON_RESULTS!$A$2:$A${DATA_ROWS+1})-1),{idx})),"")')
    ws_exc.cell(row=rr, column=3, value=f'=IF(B{rr}="","",INDEX(RECON_RESULTS!$B$2:$B${DATA_ROWS+1},SMALL(IF(RECON_RESULTS!$A$2:$A${DATA_ROWS+1}="Near Match",ROW(RECON_RESULTS!$A$2:$A${DATA_ROWS+1})-1),{idx})))')
    ws_exc.cell(row=rr, column=4, value=f'=IF(B{rr}="","",INDEX(RECON_RESULTS!$C$2:$C${DATA_ROWS+1},SMALL(IF(RECON_RESULTS!$A$2:$A${DATA_ROWS+1}="Near Match",ROW(RECON_RESULTS!$A$2:$A${DATA_ROWS+1})-1),{idx})))')
    ws_exc.cell(row=rr, column=4).number_format = DATE_FMT
    ws_exc.cell(row=rr, column=5, value=f'=IF(B{rr}="","",INDEX(RECON_RESULTS!$D$2:$D${DATA_ROWS+1},SMALL(IF(RECON_RESULTS!$A$2:$A${DATA_ROWS+1}="Near Match",ROW(RECON_RESULTS!$A$2:$A${DATA_ROWS+1})-1),{idx})))')
    ws_exc.cell(row=rr, column=5).number_format = SAR_FMT
    ws_exc.cell(row=rr, column=6, value=f'=IF(B{rr}="","",INDEX(RECON_RESULTS!$F$2:$F${DATA_ROWS+1},SMALL(IF(RECON_RESULTS!$A$2:$A${DATA_ROWS+1}="Near Match",ROW(RECON_RESULTS!$A$2:$A${DATA_ROWS+1})-1),{idx})))')
    ws_exc.cell(row=rr, column=6).number_format = DATE_FMT
    ws_exc.cell(row=rr, column=7, value=f'=IF(B{rr}="","",INDEX(RECON_RESULTS!$I$2:$I${DATA_ROWS+1},SMALL(IF(RECON_RESULTS!$A$2:$A${DATA_ROWS+1}="Near Match",ROW(RECON_RESULTS!$A$2:$A${DATA_ROWS+1})-1),{idx})))')
    ws_exc.cell(row=rr, column=7).number_format = SAR_FMT
    ws_exc.cell(row=rr, column=8, value=f'=IF(B{rr}="","",INDEX(RECON_RESULTS!$J$2:$J${DATA_ROWS+1},SMALL(IF(RECON_RESULTS!$A$2:$A${DATA_ROWS+1}="Near Match",ROW(RECON_RESULTS!$A$2:$A${DATA_ROWS+1})-1),{idx})))')
    ws_exc.cell(row=rr, column=8).number_format = NUM_FMT
    ws_exc.cell(row=rr, column=9, value=f'=IF(B{rr}="","",INDEX(RECON_RESULTS!$K$2:$K${DATA_ROWS+1},SMALL(IF(RECON_RESULTS!$A$2:$A${DATA_ROWS+1}="Near Match",ROW(RECON_RESULTS!$A$2:$A${DATA_ROWS+1})-1),{idx})))')
    ws_exc.cell(row=rr, column=10, value=f'=IF(B{rr}="","",INDEX(RECON_RESULTS!$L$2:$L${DATA_ROWS+1},SMALL(IF(RECON_RESULTS!$A$2:$A${DATA_ROWS+1}="Near Match",ROW(RECON_RESULTS!$A$2:$A${DATA_ROWS+1})-1),{idx})))')
    ws_exc.cell(row=rr, column=10).number_format = '0.0'
    ws_exc.cell(row=rr, column=11, value="")
    ws_exc.cell(row=rr, column=12, value="")

    for c in range(2, 13):
        ws_exc.cell(row=rr, column=c).border = thin_border
        ws_exc.cell(row=rr, column=c).font = data_font


# ═══════════════════════════════════════════════════════════════
# 11. EXPORT_READY SHEET
# ═══════════════════════════════════════════════════════════════
ws_exp = wb.create_sheet("EXPORT_READY")
ws_exp.sheet_properties.tabColor = P.EMERALD
ws_exp.sheet_view.showGridLines = False
set_col_widths(ws_exp, [2, 30, 40, 30])

# Instructions
exp_instructions = [
    (2, "تصدير التقارير — EXPORT & REPORTS", section_font),
    (4, "لتصدير النتائج:", Font(name='Calibri', size=11, bold=True, color=P.SLATE)),
    (5, "1. كل النتائج: انسخ جدول tbl_ReconResults من ورقة RECON_RESULTS والصقه في ملف جديد.", data_font),
    (6, "2. المطابقات فقط: استخدم الفلتر في RECON_RESULTS واختر Status = Matched.", data_font),
    (7, "3. الاستثناءات فقط: استخدم الفلتر واختر Near Match + Unmatched GL + Unmatched Bank.", data_font),
    (8, "4. أو استخدم الجداول المعدة أدناه.", data_font),
    (10, "ملخص الجلسة — SESSION SUMMARY", section_font),
]

for row_num, text, fnt in exp_instructions:
    cell = ws_exp.cell(row=row_num, column=2, value=text)
    cell.font = fnt
    cell.alignment = Alignment(horizontal='right', wrap_text=True)

# Summary stats table
summary_items = [
    ("مطابق — Matched", f'=COUNTIF(MATCH_ENGINE!N$2:N${DATA_ROWS+1},"Matched")', P.EMERALD),
    ("قريب — Near Match", f'=COUNTIF(MATCH_ENGINE!N$2:N${DATA_ROWS+1},"Near Match")', P.AMBER),
    ("GL غير مطابق — Unmatched GL", f'=COUNTIF(MATCH_ENGINE!N$2:N${DATA_ROWS+1},"Unmatched GL")', P.CRIMSON),
    ("بنك غير مطابق — Unmatched Bank", f'=COUNTIFS(BANK_CLEAN!A$2:A${DATA_ROWS+1},"<>",BANK_CLEAN!L$2:L${DATA_ROWS+1},"")-COUNTIF(BANK_CLEAN!A$2:A${DATA_ROWS+1},"")', P.AZURE),
    ("الإجمالي — Total", f'=COUNTA(MATCH_ENGINE!A$2:A${DATA_ROWS+1})-COUNTIF(MATCH_ENGINE!A$2:A${DATA_ROWS+1},"")', P.INK),
]

r = 12
ws_exp.cell(row=r, column=2, value="البند — Item").font = header_font
ws_exp.cell(row=r, column=2).fill = header_fill
ws_exp.cell(row=r, column=2).border = thin_border
ws_exp.cell(row=r, column=3, value="العدد — Count").font = header_font
ws_exp.cell(row=r, column=3).fill = header_fill
ws_exp.cell(row=r, column=3).border = thin_border

for idx, (label, formula, color) in enumerate(summary_items):
    rr = r + 1 + idx
    ws_exp.cell(row=rr, column=2, value=label).font = Font(name='Calibri', size=11, color=P.SLATE)
    ws_exp.cell(row=rr, column=2).border = thin_border
    ws_exp.cell(row=rr, column=2).fill = PatternFill(start_color=P.WHITE, end_color=P.WHITE, fill_type='solid')
    val_c = ws_exp.cell(row=rr, column=3, value=formula)
    val_c.font = Font(name='Consolas', size=14, bold=True, color=color)
    val_c.alignment = Alignment(horizontal='center')
    val_c.border = thin_border
    val_c.fill = PatternFill(start_color=P.WHITE, end_color=P.WHITE, fill_type='solid')

# VBA suggestion note
r_vba = r + len(summary_items) + 3
ws_exp.merge_cells(f'B{r_vba}:D{r_vba}')
ws_exp.cell(row=r_vba, column=2,
    value="ملاحظة: لتصدير CSV تلقائياً، يمكن إضافة ماكرو VBA اختياري. راجع ملف المشروع للتفاصيل.")
ws_exp.cell(row=r_vba, column=2).font = Font(name='Calibri', size=10, italic=True, color=P.LIGHT)

# ═══════════════════════════════════════════════════════════════
# FINAL: Save workbook
# ═══════════════════════════════════════════════════════════════
OUTPUT = "/home/user/financial-app/frontend/src/BankReconciliation.xlsx"
wb.save(OUTPUT)
print(f"Workbook saved to {OUTPUT}")
print(f"Sheets: {wb.sheetnames}")
print(f"Named ranges: {list(wb.defined_names.keys())}")

