"""
Treasury Reconciliation Command Center
Production-grade Streamlit application for corporate treasury reconciliation.
Handles SAP multi-line GL entries via Cash View reconciliation against bank statements.
"""

import streamlit as st
import pandas as pd
import numpy as np
import re
import hashlib
import io
import json
import sqlite3
import os
from datetime import datetime, timedelta
from rapidfuzz import fuzz
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------------------------------------------------------------------
# Configuration & Theme
# ---------------------------------------------------------------------------

st.set_page_config(
    page_title="Treasury Reconciliation Command Center",
    page_icon="",
    layout="wide",
    initial_sidebar_state="expanded",
)

NAVY = "#0a1628"
DARK_BG = "#0e1f3e"
CARD_BG = "#132244"
GOLD = "#c9a84c"
LIGHT_TEXT = "#e0e0e0"
GREEN = "#27ae60"
RED = "#e74c3c"
AMBER = "#f39c12"

CUSTOM_CSS = f"""
<style>
    .stApp {{
        background-color: {NAVY};
        color: {LIGHT_TEXT};
    }}
    section[data-testid="stSidebar"] {{
        background-color: {DARK_BG};
    }}
    .kpi-card {{
        background-color: {CARD_BG};
        border: 1px solid {GOLD};
        border-radius: 8px;
        padding: 18px 16px;
        text-align: center;
        margin-bottom: 8px;
    }}
    .kpi-card .value {{
        font-size: 28px;
        font-weight: 700;
        color: {GOLD};
    }}
    .kpi-card .label {{
        font-size: 13px;
        color: {LIGHT_TEXT};
        margin-top: 4px;
    }}
    .section-header {{
        color: {GOLD};
        border-bottom: 2px solid {GOLD};
        padding-bottom: 6px;
        margin-bottom: 16px;
        font-weight: 600;
    }}
    div[data-testid="stDataFrame"] {{
        border: 1px solid #1e3a5f;
        border-radius: 4px;
    }}
    .status-matched {{ color: {GREEN}; font-weight: 600; }}
    .status-exception {{ color: {RED}; font-weight: 600; }}
    .status-review {{ color: {AMBER}; font-weight: 600; }}
    .log-entry {{
        background-color: {CARD_BG};
        border-left: 3px solid {GOLD};
        padding: 6px 12px;
        margin: 4px 0;
        font-size: 13px;
    }}
</style>
"""
st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# Database / Audit helpers
# ---------------------------------------------------------------------------

DB_PATH = "treasury_audit.db"


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("""CREATE TABLE IF NOT EXISTS runs (
        run_id TEXT PRIMARY KEY, ts TEXT, user TEXT, note TEXT)""")
    conn.execute("""CREATE TABLE IF NOT EXISTS audit_events (
        event_id INTEGER PRIMARY KEY AUTOINCREMENT,
        run_id TEXT, ts TEXT, page TEXT, action TEXT,
        row_key TEXT, field TEXT, old_val TEXT, new_val TEXT, user TEXT)""")
    conn.execute("""CREATE TABLE IF NOT EXISTS overrides (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        run_id TEXT, ts TEXT, payment_key TEXT,
        field TEXT, old_val TEXT, new_val TEXT, reviewer TEXT)""")
    conn.execute("""CREATE TABLE IF NOT EXISTS exports (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        run_id TEXT, ts TEXT, export_type TEXT, filename TEXT, user TEXT)""")
    conn.commit()
    return conn


def audit_log(page, action, row_key="", field="", old_val="", new_val="", user="system"):
    run_id = st.session_state.get("run_id", "default")
    ts = datetime.utcnow().isoformat()
    entry = dict(run_id=run_id, ts=ts, page=page, action=action,
                 row_key=str(row_key), field=field,
                 old_val=str(old_val)[:500], new_val=str(new_val)[:500], user=user)
    if "audit_trail" not in st.session_state:
        st.session_state["audit_trail"] = []
    st.session_state["audit_trail"].append(entry)
    try:
        conn = get_db()
        conn.execute(
            "INSERT INTO audit_events (run_id,ts,page,action,row_key,field,old_val,new_val,user) VALUES (?,?,?,?,?,?,?,?,?)",
            (run_id, ts, page, action, str(row_key), field, str(old_val)[:500], str(new_val)[:500], user))
        conn.commit()
        conn.close()
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Session state initialization
# ---------------------------------------------------------------------------

def init_state():
    defaults = {
        "run_id": datetime.utcnow().strftime("%Y%m%d_%H%M%S"),
        "ap_raw": None, "bank_raw": None, "gl_raw": None, "manual_raw": None,
        "ap_mapped": None, "bank_mapped": None, "gl_mapped": None,
        "ap_clean": None, "bank_clean": None, "gl_clean": None,
        "cash_view": None, "match_results": None,
        "ap_col_map": {}, "bank_col_map": {}, "gl_col_map": {},
        "cash_gl_accounts": [],
        "date_tolerance": 3, "amount_tolerance": 0.01,
        "flip_bank_sign": False,
        "audit_trail": [],
        "mapping_profiles": {},
        "reviewer_name": "Treasury Analyst",
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()

# ---------------------------------------------------------------------------
# Utility functions
# ---------------------------------------------------------------------------

@st.cache_data
def load_file(uploaded_file):
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        try:
            df = pd.read_csv(uploaded_file, encoding="utf-8")
        except UnicodeDecodeError:
            uploaded_file.seek(0)
            df = pd.read_csv(uploaded_file, encoding="cp1256")
    elif name.endswith((".xlsx", ".xls")):
        df = pd.read_excel(uploaded_file, engine="openpyxl")
    else:
        st.error("Unsupported file format. Use CSV or Excel.")
        return None
    df.columns = df.columns.astype(str)
    return df


COLUMN_ALIASES = {
    "date": ["date", "valuedate", "value_date", "postingdate", "posting_date", "bookingdate",
             "booking_date", "trxdate", "transaction_date", "docdate", "doc_date",
             "تاريخ", "تاريخ_القيد"],
    "amount": ["amount", "amt", "value", "debit", "credit", "betrag", "مبلغ",
               "localamount", "local_amount", "docamount", "lc_amount", "amountinlocalcurrency"],
    "vendor": ["vendor", "vendorname", "vendor_name", "beneficiary", "beneficiaryname",
               "beneficiary_name", "payee", "name", "supplier", "المورد", "اسم_المورد",
               "name1", "vendname"],
    "iban": ["iban", "bankaccount", "bank_account", "accountno", "account_number",
             "رقم_الحساب", "bene_iban", "vendor_iban"],
    "reference": ["reference", "ref", "paymentref", "payment_ref", "paymentreference",
                   "payment_reference", "docno", "documentnumber", "document_number",
                   "المرجع", "مرجع", "belegnr"],
    "invoice": ["invoice", "invoiceno", "invoice_number", "invoiceref", "inv",
                "فاتورة", "رقم_الفاتورة", "belnr"],
    "assignment": ["assignment", "zuonr", "assign", "التخصيص"],
    "text": ["text", "description", "narrative", "details", "memo", "remark", "remarks",
             "الوصف", "النص", "sgtxt", "itemtext", "lineitemtext"],
    "currency": ["currency", "curr", "cur", "waers", "العملة"],
    "glaccount": ["glaccount", "gl_account", "account", "hkont", "glaccountno",
                  "generalledgeraccount", "حساب", "رقم_الحساب_العام"],
}


def normalize_col_name(col):
    c = col.lower().strip()
    c = re.sub(r"[\s_\-\.]+", "", c)
    return c


def auto_map_columns(df):
    mapping = {}
    norm_map = {}
    for col in df.columns:
        norm_map[normalize_col_name(col)] = col

    for std_name, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            norm_alias = re.sub(r"[\s_\-\.]+", "", alias.lower())
            if norm_alias in norm_map:
                mapping[std_name] = norm_map[norm_alias]
                break
    return mapping


def render_kpi(label, value, delta=None):
    delta_html = ""
    if delta is not None:
        color = GREEN if delta >= 0 else RED
        delta_html = f'<div style="color:{color};font-size:12px;">{delta:+.1f}%</div>'
    st.markdown(
        f'<div class="kpi-card"><div class="value">{value}</div>'
        f'<div class="label">{label}</div>{delta_html}</div>',
        unsafe_allow_html=True,
    )


# ---------------------------------------------------------------------------
# Data Cleaning
# ---------------------------------------------------------------------------

def normalize_iban(val):
    if pd.isna(val):
        return ""
    s = re.sub(r"\s+", "", str(val)).upper()
    return s


def validate_iban_length(iban):
    if not iban:
        return False
    if iban.startswith("SA") and len(iban) != 24:
        return False
    if len(iban) < 15 or len(iban) > 34:
        return False
    return True


def parse_amount(val):
    if pd.isna(val):
        return np.nan
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]
    s = s.replace(",", "").replace(" ", "")
    s = re.sub(r"[^\d.\-]", "", s)
    try:
        v = float(s)
        return -v if neg else v
    except ValueError:
        return np.nan


def parse_date(val):
    if pd.isna(val):
        return pd.NaT
    if isinstance(val, (datetime, pd.Timestamp)):
        return pd.Timestamp(val)
    s = str(val).strip()
    for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y%m%d", "%d-%m-%Y",
                "%d.%m.%Y", "%Y/%m/%d", "%d %b %Y", "%d %B %Y"]:
        try:
            return pd.Timestamp(datetime.strptime(s, fmt))
        except (ValueError, TypeError):
            continue
    try:
        return pd.Timestamp(pd.to_datetime(s, dayfirst=True))
    except Exception:
        return pd.NaT


def extract_reference(text):
    if pd.isna(text):
        return ""
    s = str(text)
    patterns = [
        r"(INV[\-_]?\d{4,20})",
        r"(PV[\-_]?\d{4,20})",
        r"(SAP[\-_]?\d{4,20})",
        r"(DOC[\-_]?\d{4,20})",
        r"(\d{6,20})",
    ]
    for pat in patterns:
        m = re.search(pat, s, re.IGNORECASE)
        if m:
            return m.group(1)
    return ""


def clean_dataframe(df, col_map, label="dataset"):
    log = {}
    out = df.copy()

    if "date" in col_map and col_map["date"] in out.columns:
        before = out[col_map["date"]].notna().sum()
        out["std_date"] = out[col_map["date"]].apply(parse_date)
        after = out["std_date"].notna().sum()
        log["Dates parsed"] = int(after)
    else:
        out["std_date"] = pd.NaT

    if "amount" in col_map and col_map["amount"] in out.columns:
        before_na = out[col_map["amount"]].isna().sum()
        out["std_amount"] = out[col_map["amount"]].apply(parse_amount)
        fixed = int((out["std_amount"].notna().sum()) - (out[col_map["amount"]].notna().sum() - before_na))
        log["Amount fixed"] = max(0, fixed)
    else:
        out["std_amount"] = np.nan

    if "iban" in col_map and col_map["iban"] in out.columns:
        out["std_iban"] = out[col_map["iban"]].apply(normalize_iban)
        log["IBAN normalized"] = int((out["std_iban"] != "").sum())
    else:
        out["std_iban"] = ""

    text_cols = []
    for k in ["text", "reference", "assignment", "invoice"]:
        if k in col_map and col_map[k] in out.columns:
            text_cols.append(col_map[k])

    if text_cols:
        out["std_text"] = out[text_cols].fillna("").astype(str).agg(" ".join, axis=1).str.strip()
    else:
        out["std_text"] = ""

    if "vendor" in col_map and col_map["vendor"] in out.columns:
        out["std_vendor"] = out[col_map["vendor"]].fillna("").astype(str).str.strip()
    else:
        out["std_vendor"] = ""

    if "reference" in col_map and col_map["reference"] in out.columns:
        out["std_ref"] = out[col_map["reference"]].fillna("").astype(str).str.strip()
    else:
        out["std_ref"] = ""

    if "invoice" in col_map and col_map["invoice"] in out.columns:
        out["std_invoice"] = out[col_map["invoice"]].fillna("").astype(str).str.strip()
    else:
        out["std_invoice"] = out["std_text"].apply(extract_reference)
        log["Invoice extracted"] = int((out["std_invoice"] != "").sum())

    if "currency" in col_map and col_map["currency"] in out.columns:
        out["std_currency"] = out[col_map["currency"]].fillna("").astype(str).str.upper().str.strip()
    else:
        out["std_currency"] = ""

    if "glaccount" in col_map and col_map["glaccount"] in out.columns:
        out["std_glaccount"] = out[col_map["glaccount"]].fillna("").astype(str).str.strip()
    else:
        out["std_glaccount"] = ""

    if "assignment" in col_map and col_map["assignment"] in out.columns:
        out["std_assignment"] = out[col_map["assignment"]].fillna("").astype(str).str.strip()
    else:
        out["std_assignment"] = ""

    empty_mask = out[["std_date", "std_amount"]].isna().all(axis=1) & (out["std_text"] == "")
    dropped = int(empty_mask.sum())
    out = out[~empty_mask].reset_index(drop=True)
    log["Dropped empty rows"] = dropped

    return out, log


# ---------------------------------------------------------------------------
# Payment Key & Cash View
# ---------------------------------------------------------------------------

def build_payment_key_gl(row):
    if row.get("std_assignment", "") and str(row["std_assignment"]).strip():
        ref = str(row["std_assignment"]).strip()
        digits = re.findall(r"\d{8,}", ref)
        if digits:
            return digits[0]
        return ref

    if row.get("std_ref", "") and str(row["std_ref"]).strip():
        ref = str(row["std_ref"]).strip()
        digits = re.findall(r"\d{8,}", ref)
        if digits:
            return digits[0]
        return ref

    txt = str(row.get("std_text", ""))
    digits = re.findall(r"\d{8,}", txt)
    if digits:
        return digits[0]

    composite = f"{row.get('std_date', '')}|{row.get('std_amount', '')}|{row.get('std_vendor', '')}"
    return "CLU_" + hashlib.md5(composite.encode()).hexdigest()[:12]


def build_payment_key_bank(row):
    if row.get("std_ref", "") and str(row["std_ref"]).strip():
        ref = str(row["std_ref"]).strip()
        digits = re.findall(r"\d{8,}", ref)
        if digits:
            return digits[0]
        return ref

    txt = str(row.get("std_text", ""))
    digits = re.findall(r"\d{8,}", txt)
    if digits:
        return digits[0]

    composite = f"{row.get('std_date', '')}|{row.get('std_amount', '')}"
    return "CLU_" + hashlib.md5(composite.encode()).hexdigest()[:12]


@st.cache_data
def build_cash_view(gl_df, cash_accounts, flip_bank_sign=False):
    if gl_df is None or gl_df.empty:
        return pd.DataFrame(), {}

    mask = gl_df["std_glaccount"].isin(cash_accounts)
    gl_cash = gl_df[mask].copy()
    gl_filtered_pct = (1 - mask.sum() / len(gl_df)) * 100 if len(gl_df) > 0 else 0

    gl_cash["payment_key"] = gl_cash.apply(build_payment_key_gl, axis=1)

    agg = gl_cash.groupby("payment_key").agg(
        gl_net_amount=("std_amount", "sum"),
        gl_min_date=("std_date", "min"),
        gl_max_date=("std_date", "max"),
        gl_line_count=("std_amount", "size"),
        gl_vendor=("std_vendor", "first"),
        gl_iban=("std_iban", "first"),
        gl_text=("std_text", "first"),
        gl_accounts=("std_glaccount", lambda x: ", ".join(x.unique())),
    ).reset_index()

    stats = {
        "total_keys": len(agg),
        "missing_keys": int(agg["payment_key"].str.startswith("CLU_").sum()),
        "gl_filtered_pct": gl_filtered_pct,
        "total_gl_lines": len(gl_cash),
    }

    return agg, stats


# ---------------------------------------------------------------------------
# Matching Engine
# ---------------------------------------------------------------------------

def compute_confidence(vendor_sim, amt_diff, date_diff_days, iban_match,
                       amount_tolerance):
    vendor_score = vendor_sim * 0.4
    if amount_tolerance > 0:
        amt_score = max(0, 40 * (1 - abs(amt_diff) / max(amount_tolerance * 10, 1)))
    else:
        amt_score = 40.0 if abs(amt_diff) < 0.01 else 0.0
    date_score = max(0, 15 * (1 - date_diff_days / 30))
    iban_bonus = 5 if iban_match else 0
    return min(100, vendor_score + amt_score + date_score + iban_bonus)


def classify_match(score):
    if score >= 90:
        return "Matched"
    elif score >= 75:
        return "Smart Match"
    elif score >= 60:
        return "Needs Review"
    else:
        return "Exception"


@st.cache_data
def match_cashview_to_bank(cash_view_df, bank_df, date_tol, amt_tol, flip_sign):
    if cash_view_df is None or cash_view_df.empty or bank_df is None or bank_df.empty:
        return pd.DataFrame()

    bank = bank_df.copy()
    bank["payment_key"] = bank.apply(build_payment_key_bank, axis=1)
    if flip_sign:
        bank["std_amount"] = -bank["std_amount"]

    results = []
    bank_used = set()
    cv = cash_view_df.copy()

    bank_key_idx = {}
    for i, row in bank.iterrows():
        bank_key_idx.setdefault(row["payment_key"], []).append(i)

    for _, gl_row in cv.iterrows():
        pk = gl_row["payment_key"]
        matched = False

        if pk in bank_key_idx and not pk.startswith("CLU_"):
            for bi in bank_key_idx[pk]:
                if bi in bank_used:
                    continue
                b_row = bank.loc[bi]
                amt_diff = abs(gl_row["gl_net_amount"] - b_row["std_amount"])
                if amt_diff <= amt_tol * max(abs(gl_row["gl_net_amount"]), 1):
                    bank_used.add(bi)
                    results.append(_build_match_row(gl_row, b_row, "KEY", 100.0, amt_diff))
                    matched = True
                    break

        if not matched:
            best_score = -1
            best_bi = None
            best_row = None
            best_diff = None
            for bi, b_row in bank.iterrows():
                if bi in bank_used:
                    continue
                amt_diff = abs(gl_row["gl_net_amount"] - b_row["std_amount"])
                if amt_diff > amt_tol * max(abs(gl_row["gl_net_amount"]), 1) + 1:
                    continue
                date_diff = 999
                if pd.notna(gl_row["gl_min_date"]) and pd.notna(b_row["std_date"]):
                    date_diff = abs((gl_row["gl_min_date"] - b_row["std_date"]).days)
                if date_diff > date_tol and date_tol > 0:
                    continue
                vendor_sim = 0
                if gl_row.get("gl_vendor", "") and b_row.get("std_vendor", ""):
                    vendor_sim = fuzz.token_set_ratio(
                        str(gl_row["gl_vendor"]), str(b_row["std_vendor"])) / 100.0
                iban_match = (gl_row.get("gl_iban", "") == b_row.get("std_iban", "")
                              and gl_row.get("gl_iban", "") != "")
                score = compute_confidence(vendor_sim, amt_diff, date_diff, iban_match, amt_tol)
                if score > best_score:
                    best_score = score
                    best_bi = bi
                    best_row = b_row
                    best_diff = amt_diff

            if best_score >= 60 and best_bi is not None:
                bank_used.add(best_bi)
                results.append(_build_match_row(gl_row, best_row, "AMT_DATE", best_score, best_diff))
                matched = True

        if not matched:
            results.append(_build_unmatched_gl(gl_row))

    for bi, b_row in bank.iterrows():
        if bi not in bank_used:
            results.append(_build_unmatched_bank(b_row))

    return pd.DataFrame(results)


def _build_match_row(gl_row, b_row, match_type, score, amt_diff):
    iban_match = (gl_row.get("gl_iban", "") == b_row.get("std_iban", "")
                  and gl_row.get("gl_iban", "") != "")
    return {
        "payment_key": gl_row["payment_key"],
        "invoice_no": "",
        "vendor_name": gl_row.get("gl_vendor", ""),
        "beneficiary_name": b_row.get("std_vendor", ""),
        "ap_amount": None,
        "gl_net_amount": gl_row["gl_net_amount"],
        "bank_amount": b_row["std_amount"],
        "bank_iban": b_row.get("std_iban", ""),
        "gl_date_range": f"{gl_row.get('gl_min_date', '')}/{gl_row.get('gl_max_date', '')}",
        "bank_date": b_row.get("std_date", ""),
        "match_type": match_type,
        "confidence_score": round(score, 1),
        "amount_diff": round(amt_diff, 2) if amt_diff is not None else None,
        "iban_match": iban_match,
        "final_status": classify_match(score),
        "reviewer": "",
        "override_flag": False,
        "notes": "",
        "audit_event_id": "",
    }


def _build_unmatched_gl(gl_row):
    return {
        "payment_key": gl_row["payment_key"],
        "invoice_no": "",
        "vendor_name": gl_row.get("gl_vendor", ""),
        "beneficiary_name": "",
        "ap_amount": None,
        "gl_net_amount": gl_row["gl_net_amount"],
        "bank_amount": None,
        "bank_iban": "",
        "gl_date_range": f"{gl_row.get('gl_min_date', '')}/{gl_row.get('gl_max_date', '')}",
        "bank_date": "",
        "match_type": "",
        "confidence_score": 0,
        "amount_diff": None,
        "iban_match": False,
        "final_status": "Pending Bank",
        "reviewer": "",
        "override_flag": False,
        "notes": "",
        "audit_event_id": "",
    }


def _build_unmatched_bank(b_row):
    return {
        "payment_key": b_row.get("payment_key", ""),
        "invoice_no": "",
        "vendor_name": "",
        "beneficiary_name": b_row.get("std_vendor", ""),
        "ap_amount": None,
        "gl_net_amount": None,
        "bank_amount": b_row["std_amount"],
        "bank_iban": b_row.get("std_iban", ""),
        "gl_date_range": "",
        "bank_date": b_row.get("std_date", ""),
        "match_type": "",
        "confidence_score": 0,
        "amount_diff": None,
        "iban_match": False,
        "final_status": "Pending GL",
        "reviewer": "",
        "override_flag": False,
        "notes": "",
        "audit_event_id": "",
    }


@st.cache_data
def match_ap_to_bank(ap_df, bank_df, date_tol, amt_tol):
    if ap_df is None or ap_df.empty or bank_df is None or bank_df.empty:
        return pd.DataFrame()

    results = []
    bank_used = set()

    for _, ap_row in ap_df.iterrows():
        best_score = -1
        best_bi = None
        best_row = None
        best_diff = None

        for bi, b_row in bank_df.iterrows():
            if bi in bank_used:
                continue

            inv_match = False
            match_type = "FUZZY"
            if ap_row.get("std_invoice", "") and b_row.get("std_text", ""):
                if str(ap_row["std_invoice"]) in str(b_row["std_text"]):
                    inv_match = True
                    match_type = "INVOICE"

            amt_diff = abs(ap_row["std_amount"] - b_row["std_amount"])
            if amt_diff > amt_tol * max(abs(ap_row["std_amount"]), 1) + 1:
                continue

            date_diff = 999
            if pd.notna(ap_row.get("std_date")) and pd.notna(b_row.get("std_date")):
                date_diff = abs((ap_row["std_date"] - b_row["std_date"]).days)
            if date_diff > date_tol and not inv_match:
                continue

            vendor_sim = 0
            if ap_row.get("std_vendor", "") and b_row.get("std_vendor", ""):
                vendor_sim = fuzz.token_set_ratio(
                    str(ap_row["std_vendor"]), str(b_row["std_vendor"])) / 100.0

            iban_match = (ap_row.get("std_iban", "") == b_row.get("std_iban", "")
                          and ap_row.get("std_iban", "") != "")

            if inv_match:
                score = 100.0
            else:
                score = compute_confidence(vendor_sim, amt_diff, date_diff, iban_match, amt_tol)

            if score > best_score:
                best_score = score
                best_bi = bi
                best_row = b_row
                best_diff = amt_diff
                if inv_match:
                    break

        if best_score >= 60 and best_bi is not None:
            bank_used.add(best_bi)
            iban_m = (ap_row.get("std_iban", "") == best_row.get("std_iban", "")
                      and ap_row.get("std_iban", "") != "")
            results.append({
                "payment_key": ap_row.get("std_ref", ""),
                "invoice_no": ap_row.get("std_invoice", ""),
                "vendor_name": ap_row.get("std_vendor", ""),
                "beneficiary_name": best_row.get("std_vendor", ""),
                "ap_amount": ap_row["std_amount"],
                "gl_net_amount": None,
                "bank_amount": best_row["std_amount"],
                "bank_iban": best_row.get("std_iban", ""),
                "gl_date_range": "",
                "bank_date": best_row.get("std_date", ""),
                "match_type": "INVOICE" if best_score == 100 else "FUZZY",
                "confidence_score": round(best_score, 1),
                "amount_diff": round(best_diff, 2),
                "iban_match": iban_m,
                "final_status": classify_match(best_score),
                "reviewer": "",
                "override_flag": False,
                "notes": "",
                "audit_event_id": "",
            })

    return pd.DataFrame(results)


# ---------------------------------------------------------------------------
# Exceptions engine
# ---------------------------------------------------------------------------

def exceptions_engine(match_df, cash_view_df, bank_df, date_tol):
    exceptions = []
    if match_df is None or match_df.empty:
        return pd.DataFrame(columns=["exception_type", "payment_key", "detail", "severity"])

    pending_bank = match_df[match_df["final_status"] == "Pending Bank"]
    for _, r in pending_bank.iterrows():
        exceptions.append({
            "exception_type": "GL not in Bank",
            "payment_key": r["payment_key"],
            "detail": f"GL amount {r.get('gl_net_amount', '')} has no bank match",
            "severity": "High",
        })

    pending_gl = match_df[match_df["final_status"] == "Pending GL"]
    for _, r in pending_gl.iterrows():
        exceptions.append({
            "exception_type": "Bank not in GL",
            "payment_key": r["payment_key"],
            "detail": f"Bank amount {r.get('bank_amount', '')} has no GL match",
            "severity": "High",
        })

    large_var = match_df[
        (match_df["amount_diff"].notna()) & (match_df["amount_diff"].abs() > 100)
    ]
    for _, r in large_var.iterrows():
        exceptions.append({
            "exception_type": "Large Variance",
            "payment_key": r["payment_key"],
            "detail": f"Amount diff: {r['amount_diff']}",
            "severity": "Medium",
        })

    iban_mismatch = match_df[
        (match_df["iban_match"] == False) &
        (match_df["bank_iban"].fillna("") != "") &
        (match_df["final_status"].isin(["Matched", "Smart Match"]))
    ]
    for _, r in iban_mismatch.iterrows():
        exceptions.append({
            "exception_type": "IBAN Mismatch",
            "payment_key": r["payment_key"],
            "detail": f"Bank IBAN: {r['bank_iban']}",
            "severity": "High",
        })

    if bank_df is not None and not bank_df.empty:
        bank_check = bank_df.copy()
        bank_check["_amt_abs"] = bank_check["std_amount"].abs()
        bank_check["_vendor_lower"] = bank_check["std_vendor"].str.lower()
        for vendor, grp in bank_check.groupby("_vendor_lower"):
            if len(grp) < 2 or not vendor:
                continue
            grp_sorted = grp.sort_values("std_date")
            for i in range(len(grp_sorted)):
                for j in range(i + 1, len(grp_sorted)):
                    ri = grp_sorted.iloc[i]
                    rj = grp_sorted.iloc[j]
                    if abs(ri["_amt_abs"] - rj["_amt_abs"]) < 0.01:
                        if pd.notna(ri["std_date"]) and pd.notna(rj["std_date"]):
                            if abs((ri["std_date"] - rj["std_date"]).days) <= date_tol:
                                exceptions.append({
                                    "exception_type": "Duplicate Payment Risk",
                                    "payment_key": f"{vendor}",
                                    "detail": f"Same vendor+amount within {date_tol} days: {ri['std_amount']}",
                                    "severity": "High",
                                })
                                break
                break

    if match_df["bank_amount"].notna().any():
        amounts = match_df["bank_amount"].dropna()
        if len(amounts) > 5:
            q1 = amounts.quantile(0.25)
            q3 = amounts.quantile(0.75)
            iqr = q3 - q1
            if iqr > 0:
                outliers = match_df[
                    (match_df["bank_amount"].notna()) &
                    ((match_df["bank_amount"] < q1 - 1.5 * iqr) |
                     (match_df["bank_amount"] > q3 + 1.5 * iqr))
                ]
                for _, r in outliers.head(20).iterrows():
                    exceptions.append({
                        "exception_type": "Outlier Payment",
                        "payment_key": r["payment_key"],
                        "detail": f"Amount {r['bank_amount']} is statistical outlier",
                        "severity": "Medium",
                    })

    return pd.DataFrame(exceptions) if exceptions else pd.DataFrame(
        columns=["exception_type", "payment_key", "detail", "severity"])


# ---------------------------------------------------------------------------
# Export functions
# ---------------------------------------------------------------------------

def export_excel(match_df, exceptions_df, cash_view_df, audit_trail, stats):
    wb = Workbook()

    header_fill = PatternFill(start_color="0A1628", end_color="0A1628", fill_type="solid")
    header_font = Font(color="C9A84C", bold=True, size=11)
    gold_border = Border(
        bottom=Side(style="thin", color="C9A84C"),
    )

    def write_df_to_sheet(ws, df, title=None):
        if title:
            ws.title = title
        if df is None or df.empty:
            ws.append(["No data available"])
            return
        for c_idx, col in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=c_idx, value=str(col))
            cell.fill = header_fill
            cell.font = header_font
            cell.border = gold_border
            cell.alignment = Alignment(horizontal="center")
        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                if isinstance(val, (pd.Timestamp, datetime)):
                    cell.value = val.strftime("%Y-%m-%d") if pd.notna(val) else ""
                elif isinstance(val, float) and np.isnan(val):
                    cell.value = ""
                else:
                    cell.value = val
        for col_cells in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 40)

    ws1 = wb.active
    if match_df is not None and not match_df.empty:
        matched = match_df[match_df["final_status"].isin(["Matched", "Smart Match"])]
        write_df_to_sheet(ws1, matched, "Matched")
    else:
        ws1.title = "Matched"
        ws1.append(["No matched records"])

    if match_df is not None and not match_df.empty:
        unmatched_bank = match_df[match_df["final_status"] == "Pending GL"]
        ws2 = wb.create_sheet("Unmatched Bank")
        write_df_to_sheet(ws2, unmatched_bank)

        unmatched_gl = match_df[match_df["final_status"] == "Pending Bank"]
        ws3 = wb.create_sheet("Unmatched GL")
        write_df_to_sheet(ws3, unmatched_gl)

    ws4 = wb.create_sheet("Exceptions")
    write_df_to_sheet(ws4, exceptions_df)

    ws5 = wb.create_sheet("Audit Log")
    if audit_trail:
        audit_df = pd.DataFrame(audit_trail)
        write_df_to_sheet(ws5, audit_df)
    else:
        ws5.append(["No audit events recorded"])

    ws6 = wb.create_sheet("Summary")
    ws6.append(["Treasury Reconciliation Summary"])
    ws6.append(["Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")])
    ws6.append([])
    if stats:
        for k, v in stats.items():
            ws6.append([k, v])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# PAGE: Upload Center
# ---------------------------------------------------------------------------

def page_upload():
    st.markdown('<h2 class="section-header">Upload Center</h2>', unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3)

    with col1:
        st.subheader("AP Payment File")
        ap_file = st.file_uploader("Upload AP file", type=["csv", "xlsx", "xls"], key="ap_upload")
        if ap_file:
            df = load_file(ap_file)
            if df is not None:
                st.session_state["ap_raw"] = df
                audit_log("Upload", "AP file uploaded", field="filename", new_val=ap_file.name)
                st.info(f"Rows: {len(df)} | Columns: {len(df.columns)}")
                st.dataframe(df.head(50), use_container_width=True, height=250)
                auto = auto_map_columns(df)
                st.session_state["ap_col_map"] = auto
                _render_column_mapper(df, "ap_col_map", "AP")

    with col2:
        st.subheader("Bank Statement File")
        bank_file = st.file_uploader("Upload Bank file", type=["csv", "xlsx", "xls"], key="bank_upload")
        if bank_file:
            df = load_file(bank_file)
            if df is not None:
                st.session_state["bank_raw"] = df
                audit_log("Upload", "Bank file uploaded", field="filename", new_val=bank_file.name)
                st.info(f"Rows: {len(df)} | Columns: {len(df.columns)}")
                st.dataframe(df.head(50), use_container_width=True, height=250)
                auto = auto_map_columns(df)
                st.session_state["bank_col_map"] = auto
                _render_column_mapper(df, "bank_col_map", "Bank")

    with col3:
        st.subheader("GL / SAP File")
        gl_file = st.file_uploader("Upload GL file", type=["csv", "xlsx", "xls"], key="gl_upload")
        if gl_file:
            df = load_file(gl_file)
            if df is not None:
                st.session_state["gl_raw"] = df
                audit_log("Upload", "GL file uploaded", field="filename", new_val=gl_file.name)
                st.info(f"Rows: {len(df)} | Columns: {len(df.columns)}")
                st.dataframe(df.head(50), use_container_width=True, height=250)
                auto = auto_map_columns(df)
                st.session_state["gl_col_map"] = auto
                _render_column_mapper(df, "gl_col_map", "GL")

    st.markdown("---")
    st.subheader("Manual SAP Verification Template (Optional)")
    manual_file = st.file_uploader("Upload Manual verification file", type=["csv", "xlsx", "xls"],
                                   key="manual_upload")
    if manual_file:
        df = load_file(manual_file)
        if df is not None:
            st.session_state["manual_raw"] = df
            st.info(f"Rows: {len(df)} | Columns: {len(df.columns)}")
            st.dataframe(df.head(50), use_container_width=True, height=200)

    st.markdown("---")
    st.subheader("Mapping Profiles")
    col_a, col_b = st.columns(2)
    with col_a:
        profile_name = st.text_input("Profile name", key="profile_name_input")
        if st.button("Save Current Mappings"):
            if profile_name:
                st.session_state["mapping_profiles"][profile_name] = {
                    "ap_col_map": st.session_state.get("ap_col_map", {}),
                    "bank_col_map": st.session_state.get("bank_col_map", {}),
                    "gl_col_map": st.session_state.get("gl_col_map", {}),
                    "cash_gl_accounts": st.session_state.get("cash_gl_accounts", []),
                }
                audit_log("Upload", "Mapping profile saved", field="profile", new_val=profile_name)
                st.success(f"Profile '{profile_name}' saved.")
    with col_b:
        profiles = list(st.session_state.get("mapping_profiles", {}).keys())
        if profiles:
            sel = st.selectbox("Load profile", [""] + profiles, key="load_profile_select")
            if sel and st.button("Load Selected Profile"):
                p = st.session_state["mapping_profiles"][sel]
                st.session_state["ap_col_map"] = p.get("ap_col_map", {})
                st.session_state["bank_col_map"] = p.get("bank_col_map", {})
                st.session_state["gl_col_map"] = p.get("gl_col_map", {})
                st.session_state["cash_gl_accounts"] = p.get("cash_gl_accounts", [])
                audit_log("Upload", "Mapping profile loaded", field="profile", new_val=sel)
                st.success(f"Profile '{sel}' loaded.")


def _render_column_mapper(df, map_key, label):
    st.markdown(f"**{label} Column Mapping**")
    current = st.session_state.get(map_key, {})
    cols_list = [""] + list(df.columns)
    for std_name in COLUMN_ALIASES.keys():
        default_idx = 0
        if std_name in current and current[std_name] in cols_list:
            default_idx = cols_list.index(current[std_name])
        sel = st.selectbox(
            f"{std_name}",
            cols_list,
            index=default_idx,
            key=f"map_{map_key}_{std_name}",
        )
        if sel:
            current[std_name] = sel
        elif std_name in current:
            del current[std_name]
    st.session_state[map_key] = current


# ---------------------------------------------------------------------------
# PAGE: Data Cleaning
# ---------------------------------------------------------------------------

def page_cleaning():
    st.markdown('<h2 class="section-header">Data Cleaning Engine</h2>', unsafe_allow_html=True)

    datasets = [
        ("AP", "ap_raw", "ap_col_map", "ap_clean"),
        ("Bank", "bank_raw", "bank_col_map", "bank_clean"),
        ("GL", "gl_raw", "gl_col_map", "gl_clean"),
    ]

    if st.button("Run Cleaning", type="primary"):
        for label, raw_key, map_key, clean_key in datasets:
            raw = st.session_state.get(raw_key)
            col_map = st.session_state.get(map_key, {})
            if raw is not None and not raw.empty:
                cleaned, log = clean_dataframe(raw, col_map, label)
                st.session_state[clean_key] = cleaned
                audit_log("Cleaning", f"{label} cleaned", field="rows", new_val=str(len(cleaned)))

    for label, raw_key, map_key, clean_key in datasets:
        raw = st.session_state.get(raw_key)
        cleaned = st.session_state.get(clean_key)

        if raw is not None:
            st.subheader(f"{label} Dataset")
            if cleaned is not None:
                _, log = clean_dataframe(raw, st.session_state.get(map_key, {}), label)
                col_a, col_b = st.columns([1, 2])
                with col_a:
                    st.markdown("**Cleaning Log**")
                    for k, v in log.items():
                        st.markdown(
                            f'<div class="log-entry">{k}: <strong>{v}</strong></div>',
                            unsafe_allow_html=True)
                with col_b:
                    std_cols = [c for c in cleaned.columns if c.startswith("std_")]
                    st.dataframe(cleaned[std_cols].head(100), use_container_width=True, height=250)
            else:
                st.info("Click 'Run Cleaning' to process this dataset.")
        else:
            pass


# ---------------------------------------------------------------------------
# PAGE: Cash View Builder
# ---------------------------------------------------------------------------

def page_cashview():
    st.markdown('<h2 class="section-header">Cash View Builder</h2>', unsafe_allow_html=True)
    st.caption("Solves the SAP 3-line GL problem by filtering to cash accounts and collapsing to net amounts.")

    gl_clean = st.session_state.get("gl_clean")
    if gl_clean is None or gl_clean.empty:
        st.warning("GL dataset not cleaned yet. Go to Data Cleaning first.")
        return

    all_accounts = sorted(gl_clean["std_glaccount"].dropna().unique().tolist())
    if not all_accounts:
        st.warning("No GL accounts found. Check GL column mapping for 'glaccount'.")
        return

    st.subheader("Configuration")
    col1, col2 = st.columns(2)

    with col1:
        selected_accounts = st.multiselect(
            "Select Cash / Bank GL Accounts",
            all_accounts,
            default=st.session_state.get("cash_gl_accounts", []),
            key="cashview_accounts",
        )
        st.session_state["cash_gl_accounts"] = selected_accounts

    with col2:
        flip = st.checkbox("Flip bank sign convention",
                           value=st.session_state.get("flip_bank_sign", False),
                           key="cashview_flip")
        st.session_state["flip_bank_sign"] = flip

        date_tol = st.number_input("Date tolerance (days)", min_value=0, max_value=30,
                                   value=st.session_state.get("date_tolerance", 3),
                                   key="cashview_date_tol")
        st.session_state["date_tolerance"] = date_tol

        amt_tol = st.number_input("Amount tolerance (fraction)", min_value=0.0, max_value=1.0,
                                  value=st.session_state.get("amount_tolerance", 0.01),
                                  step=0.005, format="%.3f", key="cashview_amt_tol")
        st.session_state["amount_tolerance"] = amt_tol

    if not selected_accounts:
        st.info("Select at least one cash GL account to build Cash View.")
        return

    if st.button("Build Cash View", type="primary"):
        audit_log("CashView", "Cash view built",
                  field="accounts", new_val=", ".join(selected_accounts))
        cv, stats = build_cash_view(gl_clean, selected_accounts, flip)
        st.session_state["cash_view"] = cv
        st.session_state["cash_view_stats"] = stats

    cv = st.session_state.get("cash_view")
    stats = st.session_state.get("cash_view_stats", {})

    if cv is not None and not cv.empty:
        st.subheader("Cash View Statistics")
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            render_kpi("Payment Keys", stats.get("total_keys", 0))
        with c2:
            render_kpi("Cluster Keys (fallback)", stats.get("missing_keys", 0))
        with c3:
            render_kpi("GL Lines (cash)", stats.get("total_gl_lines", 0))
        with c4:
            render_kpi("GL Filtered Out %", f"{stats.get('gl_filtered_pct', 0):.1f}%")

        st.subheader("Collapsed Cash View Table")
        st.dataframe(cv, use_container_width=True, height=400)

        st.subheader("Top Keys by GL Line Count")
        top_keys = cv.nlargest(20, "gl_line_count")
        st.dataframe(top_keys[["payment_key", "gl_net_amount", "gl_line_count", "gl_vendor"]],
                     use_container_width=True)

        st.subheader("Drill-Down (select a payment key)")
        key_sel = st.selectbox("Payment Key", [""] + cv["payment_key"].tolist(), key="cv_drill")
        if key_sel:
            gl_detail = gl_clean[gl_clean.apply(build_payment_key_gl, axis=1) == key_sel]
            st.dataframe(gl_detail, use_container_width=True)


# ---------------------------------------------------------------------------
# PAGE: Matching Engine
# ---------------------------------------------------------------------------

def page_matching():
    st.markdown('<h2 class="section-header">Matching Engine</h2>', unsafe_allow_html=True)

    tab_cv, tab_ap = st.tabs(["Route A: Cash View vs Bank", "Route B: AP vs Bank"])

    with tab_cv:
        cv = st.session_state.get("cash_view")
        bank_clean = st.session_state.get("bank_clean")

        if cv is None or cv.empty:
            st.warning("Cash View not built yet. Go to Cash View Builder.")
        elif bank_clean is None or bank_clean.empty:
            st.warning("Bank data not cleaned yet. Go to Data Cleaning.")
        else:
            st.info(f"Cash View keys: {len(cv)} | Bank rows: {len(bank_clean)}")
            if st.button("Run Cash View Matching", type="primary", key="run_cv_match"):
                results = match_cashview_to_bank(
                    cv, bank_clean,
                    st.session_state["date_tolerance"],
                    st.session_state["amount_tolerance"],
                    st.session_state["flip_bank_sign"],
                )
                st.session_state["match_results"] = results
                audit_log("Matching", "Cash View matching executed",
                          field="results", new_val=str(len(results)))

    with tab_ap:
        ap_clean = st.session_state.get("ap_clean")
        bank_clean = st.session_state.get("bank_clean")

        if ap_clean is None or ap_clean.empty:
            st.warning("AP data not cleaned yet.")
        elif bank_clean is None or bank_clean.empty:
            st.warning("Bank data not cleaned yet.")
        else:
            st.info(f"AP rows: {len(ap_clean)} | Bank rows: {len(bank_clean)}")
            if st.button("Run AP vs Bank Matching", type="primary", key="run_ap_match"):
                results = match_ap_to_bank(
                    ap_clean, bank_clean,
                    st.session_state["date_tolerance"],
                    st.session_state["amount_tolerance"],
                )
                if st.session_state.get("match_results") is not None:
                    st.session_state["match_results"] = pd.concat(
                        [st.session_state["match_results"], results], ignore_index=True)
                else:
                    st.session_state["match_results"] = results
                audit_log("Matching", "AP vs Bank matching executed",
                          field="results", new_val=str(len(results)))

    match_df = st.session_state.get("match_results")
    if match_df is not None and not match_df.empty:
        st.markdown("---")
        st.subheader("Match Results")

        fc1, fc2, fc3, fc4, fc5 = st.columns(5)
        with fc1:
            status_filter = st.multiselect("Status", match_df["final_status"].unique().tolist(),
                                           key="match_status_filter")
        with fc2:
            conf_range = st.slider("Confidence range", 0, 100, (0, 100), key="match_conf_filter")
        with fc3:
            vendor_filter = st.text_input("Vendor contains", key="match_vendor_filter")
        with fc4:
            diff_threshold = st.number_input("Max amount diff", min_value=0.0, value=999999.0,
                                             key="match_diff_filter")
        with fc5:
            iban_mismatch_only = st.checkbox("IBAN mismatch only", key="match_iban_filter")

        filtered = match_df.copy()
        if status_filter:
            filtered = filtered[filtered["final_status"].isin(status_filter)]
        filtered = filtered[
            (filtered["confidence_score"] >= conf_range[0]) &
            (filtered["confidence_score"] <= conf_range[1])
        ]
        if vendor_filter:
            mask = (filtered["vendor_name"].str.contains(vendor_filter, case=False, na=False) |
                    filtered["beneficiary_name"].str.contains(vendor_filter, case=False, na=False))
            filtered = filtered[mask]
        if diff_threshold < 999999:
            filtered = filtered[filtered["amount_diff"].fillna(0).abs() <= diff_threshold]
        if iban_mismatch_only:
            filtered = filtered[filtered["iban_match"] == False]

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            render_kpi("Total Results", len(filtered))
        with c2:
            matched_count = len(filtered[filtered["final_status"].isin(["Matched", "Smart Match"])])
            render_kpi("Matched", matched_count)
        with c3:
            review_count = len(filtered[filtered["final_status"] == "Needs Review"])
            render_kpi("Needs Review", review_count)
        with c4:
            exc_count = len(filtered[filtered["final_status"].isin(["Exception", "Pending Bank", "Pending GL"])])
            render_kpi("Exceptions", exc_count)

        st.dataframe(filtered, use_container_width=True, height=500)


# ---------------------------------------------------------------------------
# PAGE: Review & Approvals
# ---------------------------------------------------------------------------

def page_review():
    st.markdown('<h2 class="section-header">Review & Approvals</h2>', unsafe_allow_html=True)

    match_df = st.session_state.get("match_results")
    if match_df is None or match_df.empty:
        st.warning("No match results available. Run the Matching Engine first.")
        return

    reviewer = st.text_input("Reviewer Name",
                             value=st.session_state.get("reviewer_name", ""),
                             key="review_name_input")
    st.session_state["reviewer_name"] = reviewer

    review_items = match_df[match_df["final_status"].isin(
        ["Needs Review", "Smart Match", "Exception"])].copy()

    if review_items.empty:
        st.success("No items requiring review.")
        return

    st.info(f"{len(review_items)} items for review")

    edited = st.data_editor(
        review_items,
        column_config={
            "final_status": st.column_config.SelectboxColumn(
                "Status",
                options=["Matched", "Smart Match", "Needs Review", "Exception",
                         "Approved", "Investigate", "False Positive"],
            ),
            "override_flag": st.column_config.CheckboxColumn("Override"),
            "notes": st.column_config.TextColumn("Notes", width="large"),
            "reviewer": st.column_config.TextColumn("Reviewer"),
        },
        disabled=["payment_key", "gl_net_amount", "bank_amount", "confidence_score",
                   "match_type", "amount_diff"],
        use_container_width=True,
        height=500,
        num_rows="fixed",
        key="review_editor",
    )

    if st.button("Save Review Changes", type="primary"):
        for idx in edited.index:
            orig = review_items.loc[idx] if idx in review_items.index else None
            new = edited.loc[idx]
            if orig is not None:
                for col in ["final_status", "override_flag", "notes", "reviewer"]:
                    old_v = orig.get(col, "")
                    new_v = new.get(col, "")
                    if str(old_v) != str(new_v):
                        audit_log("Review", f"Field '{col}' changed",
                                  row_key=new.get("payment_key", ""),
                                  field=col, old_val=str(old_v), new_val=str(new_v),
                                  user=reviewer)

            if new.get("reviewer", "") == "" and reviewer:
                edited.at[idx, "reviewer"] = reviewer

        st.session_state["match_results"].update(edited)
        st.success("Review changes saved and audit log updated.")


# ---------------------------------------------------------------------------
# PAGE: Exceptions & Risk Flags
# ---------------------------------------------------------------------------

def page_exceptions():
    st.markdown('<h2 class="section-header">Exceptions & Risk Flags</h2>', unsafe_allow_html=True)

    match_df = st.session_state.get("match_results")
    cv = st.session_state.get("cash_view")
    bank_clean = st.session_state.get("bank_clean")

    if match_df is None or match_df.empty:
        st.warning("No match results available. Run the Matching Engine first.")
        return

    exc = exceptions_engine(match_df, cv, bank_clean, st.session_state.get("date_tolerance", 3))

    if exc.empty:
        st.success("No exceptions detected.")
        return

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        render_kpi("Total Exceptions", len(exc))
    with c2:
        high = len(exc[exc["severity"] == "High"])
        render_kpi("High Severity", high)
    with c3:
        render_kpi("Medium Severity", len(exc[exc["severity"] == "Medium"]))
    with c4:
        types = exc["exception_type"].nunique()
        render_kpi("Exception Types", types)

    st.subheader("Top Risk Drivers")
    risk_summary = exc["exception_type"].value_counts().reset_index()
    risk_summary.columns = ["Exception Type", "Count"]
    fig = px.bar(risk_summary, x="Exception Type", y="Count",
                 color="Count", color_continuous_scale=["#c9a84c", "#e74c3c"],
                 template="plotly_dark")
    fig.update_layout(
        paper_bgcolor=NAVY, plot_bgcolor=CARD_BG,
        font_color=LIGHT_TEXT, showlegend=False,
    )
    st.plotly_chart(fig, use_container_width=True)

    st.subheader("Exception Details")
    sev_filter = st.multiselect("Filter by severity", ["High", "Medium", "Low"],
                                default=["High", "Medium"], key="exc_sev_filter")
    type_filter = st.multiselect("Filter by type", exc["exception_type"].unique().tolist(),
                                 key="exc_type_filter")

    filtered = exc.copy()
    if sev_filter:
        filtered = filtered[filtered["severity"].isin(sev_filter)]
    if type_filter:
        filtered = filtered[filtered["exception_type"].isin(type_filter)]

    st.dataframe(filtered, use_container_width=True, height=400)


# ---------------------------------------------------------------------------
# PAGE: Executive Dashboard
# ---------------------------------------------------------------------------

def page_dashboard():
    st.markdown('<h2 class="section-header">Executive Dashboard</h2>', unsafe_allow_html=True)

    match_df = st.session_state.get("match_results")
    ap_clean = st.session_state.get("ap_clean")
    bank_clean = st.session_state.get("bank_clean")
    cv = st.session_state.get("cash_view")

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        ap_vol = ap_clean["std_amount"].abs().sum() if ap_clean is not None and not ap_clean.empty else 0
        render_kpi("Total AP Volume", f"{ap_vol:,.0f}")
    with c2:
        gl_vol = cv["gl_net_amount"].abs().sum() if cv is not None and not cv.empty else 0
        render_kpi("GL Cash Volume", f"{gl_vol:,.0f}")
    with c3:
        bank_vol = bank_clean["std_amount"].abs().sum() if bank_clean is not None and not bank_clean.empty else 0
        render_kpi("Bank Volume", f"{bank_vol:,.0f}")
    with c4:
        if match_df is not None and not match_df.empty:
            total = len(match_df)
            matched = len(match_df[match_df["final_status"].isin(["Matched", "Smart Match", "Approved"])])
            rate = (matched / total * 100) if total > 0 else 0
            render_kpi("Match Rate", f"{rate:.1f}%")
        else:
            render_kpi("Match Rate", "N/A")

    c5, c6, c7, c8 = st.columns(4)
    with c5:
        if match_df is not None and not match_df.empty:
            auto = len(match_df[match_df["match_type"].isin(["KEY", "INVOICE"])])
            auto_pct = (auto / len(match_df) * 100) if len(match_df) > 0 else 0
            render_kpi("Auto-Match %", f"{auto_pct:.1f}%")
        else:
            render_kpi("Auto-Match %", "N/A")
    with c6:
        review_count = len(match_df[match_df["final_status"] == "Needs Review"]) if match_df is not None else 0
        render_kpi("Manual Review", review_count)
    with c7:
        exc_count = len(match_df[match_df["final_status"].isin(
            ["Exception", "Pending Bank", "Pending GL"])]) if match_df is not None else 0
        render_kpi("Exceptions", exc_count)
    with c8:
        if match_df is not None and not match_df.empty:
            mismatch_exp = match_df[match_df["amount_diff"].notna()]["amount_diff"].abs().sum()
            render_kpi("Mismatch Exposure", f"{mismatch_exp:,.0f}")
        else:
            render_kpi("Mismatch Exposure", "N/A")

    if match_df is not None and not match_df.empty:
        st.markdown("---")
        col_a, col_b = st.columns(2)

        with col_a:
            status_counts = match_df["final_status"].value_counts().reset_index()
            status_counts.columns = ["Status", "Count"]
            fig1 = px.bar(status_counts, x="Status", y="Count", color="Status",
                          color_discrete_map={
                              "Matched": GREEN, "Smart Match": "#2ecc71",
                              "Needs Review": AMBER, "Exception": RED,
                              "Pending Bank": "#95a5a6", "Pending GL": "#7f8c8d",
                              "Approved": "#1abc9c",
                          },
                          template="plotly_dark",
                          title="Match Status Distribution")
            fig1.update_layout(paper_bgcolor=NAVY, plot_bgcolor=CARD_BG,
                               font_color=LIGHT_TEXT, showlegend=False)
            st.plotly_chart(fig1, use_container_width=True)

        with col_b:
            fig2 = px.histogram(match_df, x="confidence_score", nbins=20,
                                template="plotly_dark",
                                title="Confidence Score Distribution",
                                color_discrete_sequence=[GOLD])
            fig2.update_layout(paper_bgcolor=NAVY, plot_bgcolor=CARD_BG,
                               font_color=LIGHT_TEXT)
            st.plotly_chart(fig2, use_container_width=True)

        col_c, col_d = st.columns(2)

        with col_c:
            vendor_col = "vendor_name"
            if match_df[vendor_col].notna().any():
                mismatch_vendors = match_df[
                    match_df["final_status"].isin(["Needs Review", "Exception"])
                ][vendor_col].value_counts().head(10).reset_index()
                mismatch_vendors.columns = ["Vendor", "Count"]
                fig3 = px.bar(mismatch_vendors, x="Count", y="Vendor", orientation="h",
                              template="plotly_dark",
                              title="Top Mismatched Vendors",
                              color_discrete_sequence=[RED])
                fig3.update_layout(paper_bgcolor=NAVY, plot_bgcolor=CARD_BG,
                                   font_color=LIGHT_TEXT, yaxis=dict(autorange="reversed"))
                st.plotly_chart(fig3, use_container_width=True)

        with col_d:
            bank_dates = match_df[match_df["bank_date"].notna()].copy()
            if not bank_dates.empty:
                bank_dates["bank_date"] = pd.to_datetime(bank_dates["bank_date"], errors="coerce")
                daily = bank_dates.groupby(bank_dates["bank_date"].dt.date).agg(
                    bank_total=("bank_amount", lambda x: x.dropna().abs().sum()),
                    gl_total=("gl_net_amount", lambda x: x.dropna().abs().sum()),
                ).reset_index()
                daily.columns = ["Date", "Bank", "GL"]
                fig4 = go.Figure()
                fig4.add_trace(go.Scatter(x=daily["Date"], y=daily["Bank"],
                                          name="Bank", line=dict(color=GOLD)))
                fig4.add_trace(go.Scatter(x=daily["Date"], y=daily["GL"],
                                          name="GL Cash", line=dict(color=GREEN)))
                fig4.update_layout(template="plotly_dark", title="Bank vs GL Net by Day",
                                   paper_bgcolor=NAVY, plot_bgcolor=CARD_BG,
                                   font_color=LIGHT_TEXT)
                st.plotly_chart(fig4, use_container_width=True)


# ---------------------------------------------------------------------------
# PAGE: Export Center
# ---------------------------------------------------------------------------

def page_export():
    st.markdown('<h2 class="section-header">Export Center</h2>', unsafe_allow_html=True)

    match_df = st.session_state.get("match_results")
    cv = st.session_state.get("cash_view")
    bank_clean = st.session_state.get("bank_clean")
    audit_trail = st.session_state.get("audit_trail", [])

    stats = {}
    if match_df is not None and not match_df.empty:
        total = len(match_df)
        matched = len(match_df[match_df["final_status"].isin(["Matched", "Smart Match", "Approved"])])
        stats = {
            "Total Records": total,
            "Matched": matched,
            "Match Rate": f"{matched / total * 100:.1f}%" if total > 0 else "N/A",
            "Exceptions": len(match_df[match_df["final_status"].isin(
                ["Exception", "Pending Bank", "Pending GL"])]),
            "Needs Review": len(match_df[match_df["final_status"] == "Needs Review"]),
            "Timestamp": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
            "Reviewer": st.session_state.get("reviewer_name", ""),
        }

    st.subheader("Full Reconciliation Report")
    if match_df is not None and not match_df.empty:
        exc = exceptions_engine(match_df, cv, bank_clean, st.session_state.get("date_tolerance", 3))
        buf = export_excel(match_df, exc, cv, audit_trail, stats)
        st.download_button(
            "Download Full Reconciliation Excel",
            data=buf,
            file_name=f"treasury_reconciliation_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        audit_log("Export", "Full reconciliation exported", user=st.session_state.get("reviewer_name", "system"))
    else:
        st.info("No match results to export. Run the Matching Engine first.")

    st.markdown("---")
    st.subheader("Exceptions Only Report")
    if match_df is not None and not match_df.empty:
        exc = exceptions_engine(match_df, cv, bank_clean, st.session_state.get("date_tolerance", 3))
        if not exc.empty:
            exc_buf = export_excel(pd.DataFrame(), exc, None, [], {"Report": "Exceptions Only"})
            st.download_button(
                "Download Exceptions Report",
                data=exc_buf,
                file_name=f"exceptions_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="exc_download",
            )
        else:
            st.success("No exceptions to export.")

    st.markdown("---")
    st.subheader("Audit Trail Export")
    if audit_trail:
        audit_df = pd.DataFrame(audit_trail)
        csv_buf = audit_df.to_csv(index=False).encode()
        st.download_button(
            "Download Audit Trail CSV",
            data=csv_buf,
            file_name=f"audit_trail_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv",
            key="audit_csv_download",
        )
    else:
        st.info("No audit events recorded yet.")

    st.markdown("---")
    st.subheader("Export Configuration Snapshot")
    config = {
        "ap_col_map": st.session_state.get("ap_col_map", {}),
        "bank_col_map": st.session_state.get("bank_col_map", {}),
        "gl_col_map": st.session_state.get("gl_col_map", {}),
        "cash_gl_accounts": st.session_state.get("cash_gl_accounts", []),
        "date_tolerance": st.session_state.get("date_tolerance", 3),
        "amount_tolerance": st.session_state.get("amount_tolerance", 0.01),
        "flip_bank_sign": st.session_state.get("flip_bank_sign", False),
        "timestamp": datetime.utcnow().isoformat(),
    }
    config_json = json.dumps(config, indent=2, default=str)
    st.download_button(
        "Download Configuration JSON",
        data=config_json,
        file_name="reconciliation_config.json",
        mime="application/json",
        key="config_download",
    )


# ---------------------------------------------------------------------------
# PAGE: Audit Mode
# ---------------------------------------------------------------------------

def page_audit():
    st.markdown('<h2 class="section-header">Audit Mode</h2>', unsafe_allow_html=True)

    audit_trail = st.session_state.get("audit_trail", [])

    c1, c2, c3 = st.columns(3)
    with c1:
        render_kpi("Total Audit Events", len(audit_trail))
    with c2:
        pages = set(e.get("page", "") for e in audit_trail)
        render_kpi("Pages Tracked", len(pages))
    with c3:
        users = set(e.get("user", "") for e in audit_trail)
        render_kpi("Users", len(users))

    st.subheader("Session Audit Trail")
    if audit_trail:
        audit_df = pd.DataFrame(audit_trail)
        page_filter = st.multiselect("Filter by page", audit_df["page"].unique().tolist(),
                                     key="audit_page_filter")
        action_filter = st.text_input("Filter by action", key="audit_action_filter")

        filtered = audit_df.copy()
        if page_filter:
            filtered = filtered[filtered["page"].isin(page_filter)]
        if action_filter:
            filtered = filtered[filtered["action"].str.contains(action_filter, case=False, na=False)]

        st.dataframe(filtered.sort_values("ts", ascending=False),
                     use_container_width=True, height=400)
    else:
        st.info("No audit events recorded in this session.")

    st.markdown("---")
    st.subheader("Database Audit Records")
    try:
        conn = get_db()
        db_events = pd.read_sql("SELECT * FROM audit_events ORDER BY ts DESC LIMIT 500", conn)
        conn.close()
        if not db_events.empty:
            st.dataframe(db_events, use_container_width=True, height=300)
        else:
            st.info("No records in database yet.")
    except Exception as e:
        st.info(f"Database not available: {e}")

    st.subheader("Run History")
    try:
        conn = get_db()
        runs = pd.read_sql("SELECT * FROM runs ORDER BY ts DESC LIMIT 50", conn)
        conn.close()
        if not runs.empty:
            st.dataframe(runs, use_container_width=True)
        else:
            st.info("No previous runs recorded.")
    except Exception:
        st.info("Database not available.")


# ---------------------------------------------------------------------------
# MAIN: Sidebar Navigation
# ---------------------------------------------------------------------------

def main():
    try:
        conn = get_db()
        run_id = st.session_state.get("run_id", "default")
        conn.execute("INSERT OR IGNORE INTO runs (run_id, ts, user, note) VALUES (?, ?, ?, ?)",
                     (run_id, datetime.utcnow().isoformat(), "system", "session start"))
        conn.commit()
        conn.close()
    except Exception:
        pass

    st.sidebar.markdown(
        f'<h1 style="color:{GOLD};font-size:22px;margin-bottom:0;">Treasury Reconciliation</h1>'
        f'<p style="color:{LIGHT_TEXT};font-size:12px;margin-top:0;">Command Center</p>',
        unsafe_allow_html=True,
    )
    st.sidebar.markdown("---")

    pages = {
        "Upload Center": page_upload,
        "Data Cleaning": page_cleaning,
        "Cash View Builder": page_cashview,
        "Matching Engine": page_matching,
        "Review & Approvals": page_review,
        "Exceptions & Risk Flags": page_exceptions,
        "Executive Dashboard": page_dashboard,
        "Export Center": page_export,
        "Audit Mode": page_audit,
    }

    selection = st.sidebar.radio("Navigation", list(pages.keys()), label_visibility="collapsed")

    st.sidebar.markdown("---")
    st.sidebar.markdown(f'<p style="color:{LIGHT_TEXT};font-size:11px;">Run ID: {st.session_state.get("run_id", "")}</p>',
                        unsafe_allow_html=True)

    datasets = {"AP": st.session_state.get("ap_raw"),
                "Bank": st.session_state.get("bank_raw"),
                "GL": st.session_state.get("gl_raw")}
    for name, df in datasets.items():
        if df is not None:
            st.sidebar.markdown(
                f'<p style="color:{GREEN};font-size:11px;">{name}: {len(df)} rows</p>',
                unsafe_allow_html=True)
        else:
            st.sidebar.markdown(
                f'<p style="color:#666;font-size:11px;">{name}: not loaded</p>',
                unsafe_allow_html=True)

    pages[selection]()


if __name__ == "__main__":
    main()
